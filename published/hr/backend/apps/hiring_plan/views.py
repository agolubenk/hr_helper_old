"""
Документация по проблемным импортам (линтер):
- django.shortcuts, django.contrib.*, django.http, django.views.*, django.urls, django.db, django.utils

Влияние: все представления и AJAX-эндпоинты модуля `hiring_plan` (списки/детали/формы планов,
SLA, KPI/OKR, метрики, годовой план) завязаны на эти импорты. Если импорты реально не доступны
в окружении, весь UI и API этого приложения перестанут работать.
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from django.urls import reverse_lazy, reverse
from django.db import models
from django.db.models import Q, Count, Sum, Avg
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import timedelta, date
import json

from .forms import HiringRequestForm, HiringRequestUpdateForm, VacancySLAForm, HiringPlanFilterForm

from .models import (
    HiringPlan, HiringPlanPosition, PositionType, PlanPeriodType,
    PositionKPIOKR, PlanKPIOKRBlock, PlanMetrics,
    VacancySLA, HiringRequest, RecruitmentMetrics, DemandForecast, RecruiterCapacity
)
from .services import HiringPlanServiceExtended
from .metrics_service import MetricsService


class HiringPlanListView(LoginRequiredMixin, ListView):
    """Список планов найма"""
    model = HiringPlan
    template_name = 'hiring_plan/plan_list.html'
    context_object_name = 'plans'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = HiringPlan.objects.select_related('period_type', 'owner').prefetch_related('positions')
        
        # Фильтрация
        search = self.request.GET.get('search')
        period_type = self.request.GET.get('period_type')
        is_completed = self.request.GET.get('is_completed')
        owner = self.request.GET.get('owner')
        
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        
        if period_type:
            queryset = queryset.filter(period_type_id=period_type)
        
        if is_completed == 'active':
            queryset = queryset.filter(is_completed=False)
        elif is_completed == 'completed':
            queryset = queryset.filter(is_completed=True)
        
        if owner:
            queryset = queryset.filter(owner_id=owner)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = HiringPlanFilterForm(self.request.GET)
        context['total_plans'] = HiringPlan.objects.count()
        context['active_plans'] = HiringPlan.objects.filter(is_completed=False).count()
        context['completed_plans'] = HiringPlan.objects.filter(is_completed=True).count()
        return context


class HiringPlanDetailView(LoginRequiredMixin, DetailView):
    """Детальный просмотр плана найма"""
    model = HiringPlan
    template_name = 'hiring_plan/plan_detail.html'
    context_object_name = 'plan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plan = self.get_object()
        
        # Позиции плана
        positions = plan.positions.select_related('vacancy', 'position_type').prefetch_related('grades')
        context['positions'] = positions
        
        # Статистика
        context['total_positions'] = positions.count()
        context['total_headcount_needed'] = positions.aggregate(
            total=Sum('headcount_needed'))['total'] or 0
        context['total_headcount_hired'] = positions.aggregate(
            total=Sum('headcount_hired'))['total'] or 0
        
        # Статистика по типам позиций
        position_type_stats = positions.values('position_type__name').annotate(
            count=Count('id'),
            headcount_needed=Sum('headcount_needed'),
            headcount_hired=Sum('headcount_hired')
        )
        context['position_type_stats'] = position_type_stats
        
        # Статистика по грейдам
        grade_stats = positions.values('grades__name').annotate(
            count=Count('id'),
            headcount_needed=Sum('headcount_needed'),
            headcount_hired=Sum('headcount_hired')
        ).filter(grades__isnull=False)
        context['grade_stats'] = grade_stats
        
        # SLA compliance
        sla_compliance = HiringPlanServiceExtended.get_plan_sla_compliance(plan)
        context['sla_compliance'] = sla_compliance
        
        # KPI/OKR summary
        kpi_okr_summary = HiringPlanServiceExtended.get_kpi_okr_summary(plan)
        context['kpi_okr_summary'] = kpi_okr_summary
        
        return context


class HiringPlanCreateView(LoginRequiredMixin, CreateView):
    """Создание плана найма"""
    model = HiringPlan
    # form_class = HiringPlanFormExtended  # Удалено - форма не существует
    template_name = 'hiring_plan/plan_form.html'
    success_url = reverse_lazy('hiring_plan:plan_list')
    
    def form_valid(self, form):
        form.instance.owner = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'План "{form.instance.title}" успешно создан!')
        return response


class HiringPlanUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование плана найма"""
    model = HiringPlan
    # form_class = HiringPlanFormExtended  # Удалено - форма не существует
    template_name = 'hiring_plan/plan_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'План "{form.instance.title}" успешно обновлен!')
        return response


class HiringPlanDeleteView(LoginRequiredMixin, DeleteView):
    """Удаление плана найма"""
    model = HiringPlan
    template_name = 'hiring_plan/plan_confirm_delete.html'
    success_url = reverse_lazy('hiring_plan:plan_list')
    
    def delete(self, request, *args, **kwargs):
        plan = self.get_object()
        messages.success(request, f'План "{plan.title}" успешно удален!')
        return super().delete(request, *args, **kwargs)


class PeriodicPlanCreateView(LoginRequiredMixin, CreateView):
    """Создание периодического плана"""
    template_name = 'hiring_plan/periodic_plan_create.html'
    # form_class = PeriodPlanCreationForm  # Удалено - форма не существует
    success_url = reverse_lazy('hiring_plan:plan_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['period_types'] = PlanPeriodType.objects.filter(is_active=True)
        return context
    
    def form_valid(self, form):
        # Создаем план с помощью сервиса
        plan = HiringPlanServiceExtended.create_periodic_plan(
            title=form.cleaned_data['title'],
            period_type=form.cleaned_data['period_type'],
            description=form.cleaned_data.get('description', ''),
            owner=self.request.user
        )
        
        messages.success(self.request, f'Периодический план "{plan.title}" успешно создан!')
        return redirect('hiring_plan:plan_detail', pk=plan.pk)


    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['plan_pk'] = self.kwargs['plan_pk']
        return kwargs
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'SLA для "{form.instance.vacancy.name}" успешно создано!')
        return response


class PlanSLAComplianceView(LoginRequiredMixin, DetailView):
    """SLA compliance для плана"""
    model = HiringPlan
    template_name = 'hiring_plan/plan_sla_compliance.html'
    context_object_name = 'plan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plan = self.get_object()
        
        # SLA compliance summary
        sla_compliance = HiringPlanServiceExtended.get_plan_sla_compliance(plan)
        context['sla_compliance'] = sla_compliance
        
        # Position type statistics
        position_type_stats = HiringPlanServiceExtended.get_position_type_statistics(plan)
        context['position_type_stats'] = position_type_stats
        
        # Replacement reasons statistics
        replacement_stats = HiringPlanServiceExtended.get_replacement_reasons_stats(plan)
        context['replacement_stats'] = replacement_stats
        
        return context


class PlanKPIOKRDashboardView(LoginRequiredMixin, DetailView):
    """KPI/OKR dashboard для плана"""
    model = HiringPlan
    template_name = 'hiring_plan/plan_kpi_okr_dashboard.html'
    context_object_name = 'plan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plan = self.get_object()
        
        # KPI/OKR summary
        kpi_okr_summary = HiringPlanServiceExtended.get_kpi_okr_summary(plan)
        context['kpi_okr_summary'] = kpi_okr_summary
        
        # KPI/OKR vs SLA comparison
        kpi_okr_vs_sla = HiringPlanServiceExtended.compare_plan_with_kpi_okr(plan)
        context['kpi_okr_vs_sla'] = kpi_okr_vs_sla
        
        return context


class PlanKPIOKRBlockListView(LoginRequiredMixin, ListView):
    """Список блоков KPI/OKR"""
    model = PlanKPIOKRBlock
    template_name = 'hiring_plan/kpi_okr_block_list.html'
    context_object_name = 'blocks'
    paginate_by = 20
    
    def get_queryset(self):
        return PlanKPIOKRBlock.objects.prefetch_related('position_types', 'grades').order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hiring_plans'] = HiringPlan.objects.filter(is_completed=False).order_by('-created_at')
        return context


class PlanKPIOKRBlockCreateView(LoginRequiredMixin, CreateView):
    """Создание блока KPI/OKR"""
    model = PlanKPIOKRBlock
    # form_class = PlanKPIOKRBlockForm  # Удалено - форма не существует
    template_name = 'hiring_plan/kpi_okr_block_form.html'
    success_url = reverse_lazy('hiring_plan:kpi_okr_block_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Блок KPI/OKR "{form.instance.name}" успешно создан!')
        return response


class PlanKPIOKRBlockUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование блока KPI/OKR"""
    model = PlanKPIOKRBlock
    # form_class = PlanKPIOKRBlockForm  # Удалено - форма не существует
    template_name = 'hiring_plan/kpi_okr_block_form.html'
    success_url = reverse_lazy('hiring_plan:kpi_okr_block_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Блок KPI/OKR "{form.instance.name}" успешно обновлен!')
        return response


class PositionKPIOKRCreateView(LoginRequiredMixin, CreateView):
    """Создание KPI/OKR"""
    model = PositionKPIOKR
    # form_class = PositionKPIOKRForm  # Удалено - форма не существует
    template_name = 'hiring_plan/kpi_okr_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.kwargs['plan_pk']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plan'] = get_object_or_404(HiringPlan, pk=self.kwargs['plan_pk'])
        return context
    
    def form_valid(self, form):
        form.instance.hiring_plan = get_object_or_404(HiringPlan, pk=self.kwargs['plan_pk'])
        response = super().form_valid(form)
        messages.success(self.request, f'KPI/OKR "{form.instance.name}" успешно создан!')
        return response


class PositionKPIOKRUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование KPI/OKR"""
    model = PositionKPIOKR
    # form_class = PositionKPIOKRForm  # Удалено - форма не существует
    template_name = 'hiring_plan/kpi_okr_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.object.hiring_plan.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'KPI/OKR "{form.instance.name}" успешно обновлен!')
        return response


class HiringPlanPositionCreateView(LoginRequiredMixin, CreateView):
    """Создание позиции в плане"""
    model = HiringPlanPosition
    # form_class = HiringPlanPositionFormExtended  # Удалено - форма не существует
    template_name = 'hiring_plan/position_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.kwargs['plan_pk']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['plan'] = get_object_or_404(HiringPlan, pk=self.kwargs['plan_pk'])
        return context
    
    def form_valid(self, form):
        form.instance.hiring_plan = get_object_or_404(HiringPlan, pk=self.kwargs['plan_pk'])
        response = super().form_valid(form)
        messages.success(self.request, f'Позиция "{form.instance.vacancy.name}" успешно добавлена в план!')
        return response


class HiringPlanPositionUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование позиции в плане"""
    model = HiringPlanPosition
    # form_class = HiringPlanPositionFormExtended  # Удалено - форма не существует
    template_name = 'hiring_plan/position_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.object.hiring_plan.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Позиция "{form.instance.vacancy.name}" успешно обновлена!')
        return response


class HiringPlanPositionDeleteView(LoginRequiredMixin, DeleteView):
    """Удаление позиции из плана"""
    model = HiringPlanPosition
    template_name = 'hiring_plan/position_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:plan_detail', kwargs={'pk': self.object.hiring_plan.pk})
    
    def delete(self, request, *args, **kwargs):
        position = self.get_object()
        messages.success(request, f'Позиция "{position.vacancy.name}" успешно удалена из плана!')
        return super().delete(request, *args, **kwargs)


@login_required
@require_http_methods(["POST"])
def auto_move_unfilled_positions(request, pk):
    """Автоматическое перемещение незакрытых позиций в следующий период"""
    plan = get_object_or_404(HiringPlan, pk=pk)
    
    try:
        moved_count = HiringPlanServiceExtended.auto_move_unfilled_positions(plan)
        return JsonResponse({
            'success': True,
            'message': f'Успешно перемещено {moved_count} позиций в следующий период'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def plan_ajax_data(request, pk):
    """AJAX данные для плана (для графиков и статистики)"""
    plan = get_object_or_404(HiringPlan, pk=pk)
    
    # Данные для графиков
    data = {
        'completion_rate': plan.completion_rate,
        'total_positions': plan.total_positions,
        'total_headcount_needed': plan.total_headcount_needed,
        'total_headcount_hired': plan.total_headcount_hired,
        'position_type_stats': list(plan.positions.values('position_type__name').annotate(
            count=Count('id'),
            headcount_needed=Sum('headcount_needed'),
            headcount_hired=Sum('headcount_hired')
        )),
        'grade_stats': list(plan.positions.values('grades__name').annotate(
            count=Count('id'),
            headcount_needed=Sum('headcount_needed'),
            headcount_hired=Sum('headcount_hired')
        ).filter(grades__isnull=False)),
    }
    
    return JsonResponse(data)


@login_required
@require_http_methods(["POST"])
def update_position_headcount(request, pk):
    """Обновление количества нанятых для позиции"""
    position = get_object_or_404(HiringPlanPosition, pk=pk)
    
    try:
        data = json.loads(request.body)
        headcount_hired = int(data.get('headcount_hired', 0))
        
        if headcount_hired < 0:
            return JsonResponse({'success': False, 'error': 'Количество не может быть отрицательным'})
        
        if headcount_hired > position.headcount_needed:
            return JsonResponse({'success': False, 'error': 'Количество нанятых не может превышать требуемое'})
        
        position.headcount_hired = headcount_hired
        position.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Количество нанятых успешно обновлено',
            'fulfillment_rate': position.fulfillment_rate,
            'is_fulfilled': position.is_fulfilled
        })
        
    except (ValueError, KeyError) as e:
        return JsonResponse({'success': False, 'error': 'Неверные данные'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def apply_kpi_okr_block_to_plan(request, plan_pk, block_pk):
    """Применение блока KPI/OKR к плану"""
    plan = get_object_or_404(HiringPlan, pk=plan_pk)
    block = get_object_or_404(PlanKPIOKRBlock, pk=block_pk)
    
    try:
        HiringPlanServiceExtended.apply_kpi_okr_block_to_plan(block, plan)
        messages.success(request, f'Блок KPI/OKR "{block.name}" успешно применен к плану!')
        return redirect('hiring_plan:plan_detail', pk=plan.pk)
    except Exception as e:
        messages.error(request, f'Ошибка при применении блока: {str(e)}')
        return redirect('hiring_plan:plan_detail', pk=plan.pk)


class HiringRequestsListView(LoginRequiredMixin, ListView):
    """Единый список всех заявок на найм"""
    model = HiringRequest
    template_name = 'hiring_plan/hiring_requests_list.html'
    context_object_name = 'requests'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = HiringRequest.objects.select_related(
            'vacancy', 'grade', 'sla', 'created_by', 'closed_by'
        ).order_by('-opening_date', 'priority')
        
        # Фильтры
        status = self.request.GET.get('status')
        period = self.request.GET.get('period')
        grade = self.request.GET.get('grade')
        vacancy = self.request.GET.get('vacancy')
        priority = self.request.GET.get('priority')
        opening_reason = self.request.GET.get('opening_reason')
        recruiter = self.request.GET.get('recruiter')
        search = self.request.GET.get('search')
        
        if status:
            queryset = queryset.filter(status=status)
        
        if period:
            # Фильтрация по месяцу
            try:
                year, month = period.split('-')
                queryset = queryset.filter(opening_date__year=year, opening_date__month=month)
            except ValueError:
                pass
        
        if grade:
            queryset = queryset.filter(grade_id=grade)
        
        if vacancy:
            queryset = queryset.filter(vacancy_id=vacancy)
        
        if priority:
            queryset = queryset.filter(priority=priority)
        
        if opening_reason:
            queryset = queryset.filter(opening_reason=opening_reason)
        
        if recruiter:
            queryset = queryset.filter(closed_by_id=recruiter)
        
        if search:
            queryset = queryset.filter(
                Q(vacancy__name__icontains=search) |
                Q(candidate_name__icontains=search) |
                Q(candidate_id__icontains=search) |
                Q(notes__icontains=search) |
                Q(project__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика
        requests = self.get_queryset()
        context['total_requests'] = requests.count()
        context['planned_requests'] = requests.filter(status='planned').count()
        context['in_progress_requests'] = requests.filter(status='in_progress').count()
        context['closed_requests'] = requests.filter(status='closed').count()
        context['cancelled_requests'] = requests.filter(status='cancelled').count()
        
        # Опции для фильтров
        from apps.company_settings.utils import get_active_grades_queryset
        active_grades = get_active_grades_queryset()
        
        context['status_choices'] = HiringRequest.STATUS_CHOICES
        context['priority_choices'] = HiringRequest.PRIORITY_CHOICES
        context['reason_choices'] = HiringRequest.REASON_CHOICES
        # Фильтруем только активные грейды компании для фильтров
        context['grade_choices'] = HiringRequest.objects.filter(
            grade__in=active_grades
        ).values_list('grade__id', 'grade__name').distinct()
        context['vacancy_choices'] = HiringRequest.objects.values_list('vacancy__id', 'vacancy__name').distinct()
        context['recruiter_choices'] = HiringRequest.objects.filter(
            closed_by__isnull=False
        ).values_list('closed_by__id', 'closed_by__username').distinct().order_by('closed_by__username')
        
        return context


class HiringRequestDetailView(LoginRequiredMixin, DetailView):
    """Детальный просмотр заявки"""
    model = HiringRequest
    template_name = 'hiring_plan/hiring_request_detail.html'
    context_object_name = 'hiring_request'


def _fetch_clickup_task_into_notes(hiring_request, request):
    """Если у заявки указан clickup_task_id и у пользователя есть API-ключ, подтягивает данные задачи в заметки."""
    import logging
    import re
    raw = (hiring_request.clickup_task_id or '').strip()
    task_id = raw
    # Извлекаем ID из URL, если вставлена ссылка (например https://app.clickup.com/t/86c7y88xk)
    if task_id and ('clickup.com' in task_id or '/t/' in task_id):
        m = re.search(r'/t/([a-zA-Z0-9_-]+)', task_id)
        if m:
            task_id = m.group(1)
            if task_id != raw:
                hiring_request.clickup_task_id = task_id
                hiring_request.save(update_fields=['clickup_task_id'])
    if not task_id:
        return
    api_key = getattr(request.user, 'clickup_api_key', None) or ''
    if not (api_key and api_key.strip()):
        messages.warning(
            request,
            'Укажите API-токен ClickUp в профиле, чтобы подгружать данные задачи в заметки.'
        )
        return
    try:
        from apps.clickup_int.services import ClickUpService, ClickUpAPIError
        from apps.clickup_int.hiring_plan_sync import format_clickup_task_for_notes
        service = ClickUpService(api_key.strip())
        task_data = service.get_task(task_id)
        if task_data:
            new_block = format_clickup_task_for_notes(task_data)
            current_notes = hiring_request.notes or ''
            marker_start = "--- Данные из ClickUp"
            marker_end = "--- Конец данных ClickUp ---"
            if marker_start in current_notes and marker_end in current_notes:
                start_i = current_notes.find(marker_start)
                end_i = current_notes.find(marker_end) + len(marker_end)
                before = current_notes[:start_i].rstrip()
                after = current_notes[end_i:].lstrip()
                hiring_request.notes = "\n\n".join(filter(None, [before, new_block, after]))
            else:
                hiring_request.notes = (current_notes.rstrip() + "\n\n" + new_block).strip()
            hiring_request.save(update_fields=['notes'])
            messages.info(request, 'Данные из задачи ClickUp добавлены в заметки.')
        else:
            messages.warning(request, 'Задача в ClickUp не найдена. Проверьте ID задачи.')
    except ClickUpAPIError as e:
        logging.getLogger(__name__).warning('ClickUp при подтягивании в заметки: %s', e)
        messages.warning(request, f'ClickUp: {str(e)}')
    except Exception as e:
        logging.getLogger(__name__).exception('Ошибка подтягивания данных ClickUp в заметки')
        messages.warning(request, f'Не удалось загрузить данные из ClickUp: {str(e)}')


class HiringRequestCreateView(LoginRequiredMixin, CreateView):
    """Создание заявки"""
    model = HiringRequest
    form_class = HiringRequestForm
    template_name = 'hiring_plan/hiring_request_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:hiring_requests_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Если назначен рекрутер, создаем запись в истории назначений
        if form.instance.recruiter:
            form.instance.assign_recruiter(form.instance.recruiter, self.request.user)
        
        # Синхронизируем рекрутера с вакансией
        self.object.sync_recruiter_with_vacancy()
        
        # Подтягиваем данные из задачи ClickUp в заметки, если указан clickup_task_id
        _fetch_clickup_task_into_notes(self.object, self.request)
        
        # Если есть candidate_id, пытаемся получить данные кандидата из Huntflow
        if self.object.candidate_id:
            success = self.object.fetch_candidate_data_from_huntflow(self.request.user)
            if success:
                updated_info = []
                if self.object.candidate_name:
                    updated_info.append(f'имя: {self.object.candidate_name}')
                if self.object.closed_date:
                    updated_info.append(f'дата принятия оффера: {self.object.closed_date.strftime("%d.%m.%Y")}')
                if self.object.hire_date:
                    updated_info.append(f'дата выхода: {self.object.hire_date.strftime("%d.%m.%Y")}')
                
                if updated_info:
                    messages.info(self.request, f'Данные кандидата автоматически получены из Huntflow: {", ".join(updated_info)}')
                else:
                    messages.info(self.request, 'Данные кандидата проверены в Huntflow, но новых данных не найдено')
            else:
                messages.warning(self.request, 'Не удалось получить данные кандидата из Huntflow. Проверьте ID кандидата.')
        
        messages.success(self.request, f'Заявка "{self.object}" успешно создана!')
        return response


class HiringRequestUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование заявки"""
    model = HiringRequest
    form_class = HiringRequestUpdateForm
    template_name = 'hiring_plan/hiring_request_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:hiring_requests_list')
    
    def form_valid(self, form):
        # Получаем старые значения из базы данных для точного сравнения
        old_request = HiringRequest.objects.get(id=self.object.id)
        old_opening_date = old_request.opening_date
        old_recruiter = old_request.recruiter
        
        # Сохраняем форму (это обновит self.object)
        response = super().form_valid(form)
        
        # Для незапланированных заявок сохраняем исходную дату открытия
        if self.object.status != 'planned' and self.object.opening_date:
            form.instance.opening_date = self.object.opening_date
        
        # Если изменилась дата открытия, пересчитываем статус
        if old_opening_date != self.object.opening_date:
            self.object.save()  # Это вызовет post_save сигнал, который пересчитает статус
        
        # Если изменился рекрутер, обновляем назначение
        old_recruiter_id = old_recruiter.id if old_recruiter else None
        new_recruiter_id = form.instance.recruiter.id if form.instance.recruiter else None
        
        if old_recruiter_id != new_recruiter_id:
            if form.instance.recruiter:
                self.object.assign_recruiter(form.instance.recruiter, self.request.user)
            else:
                self.object.unassign_recruiter(self.request.user)
            self.object.sync_recruiter_with_vacancy()
        
        # Если есть candidate_id, пытаемся получить данные кандидата из Huntflow
        if self.object.candidate_id:
            success = self.object.fetch_candidate_data_from_huntflow(self.request.user)
            if success:
                updated_info = []
                if self.object.candidate_name:
                    updated_info.append(f'имя: {self.object.candidate_name}')
                if self.object.closed_date:
                    updated_info.append(f'дата принятия оффера: {self.object.closed_date.strftime("%d.%m.%Y")}')
                if self.object.hire_date:
                    updated_info.append(f'дата выхода: {self.object.hire_date.strftime("%d.%m.%Y")}')
                
                if updated_info:
                    messages.info(self.request, f'Данные кандидата автоматически получены из Huntflow: {", ".join(updated_info)}')
                else:
                    messages.info(self.request, 'Данные кандидата проверены в Huntflow, но новых данных не найдено')
            else:
                messages.warning(self.request, 'Не удалось получить данные кандидата из Huntflow. Проверьте ID кандидата.')

        # Если указана связь с задачей ClickUp — подтягиваем данные задачи в заметки
        _fetch_clickup_task_into_notes(self.object, self.request)
        
        messages.success(self.request, f'Заявка "{self.object}" успешно обновлена!')
        return response


# SLA Views
class VacancySLAListView(LoginRequiredMixin, ListView):
    """Список SLA для вакансий"""
    model = VacancySLA
    template_name = 'hiring_plan/sla_list.html'
    context_object_name = 'slas'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = VacancySLA.objects.select_related('vacancy', 'grade').order_by('vacancy__name', 'grade__name')
        
        # Фильтр по вакансии
        vacancy_id = self.request.GET.get('vacancy')
        if vacancy_id:
            queryset = queryset.filter(vacancy_id=vacancy_id)
        
        # Фильтр по грейду
        grade_id = self.request.GET.get('grade')
        if grade_id:
            queryset = queryset.filter(grade_id=grade_id)
        
        # Фильтр по активности
        is_active = self.request.GET.get('is_active')
        if is_active in ['true', 'false']:
            queryset = queryset.filter(is_active=is_active == 'true')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем списки для фильтров
        from apps.vacancies.models import Vacancy
        from apps.company_settings.utils import get_active_grades_queryset
        
        context['vacancies'] = Vacancy.objects.filter(is_active=True).order_by('name')
        context['grades'] = get_active_grades_queryset().order_by('name')
        
        # Добавляем текущие значения фильтров
        context['current_vacancy'] = self.request.GET.get('vacancy', '')
        context['current_grade'] = self.request.GET.get('grade', '')
        context['current_is_active'] = self.request.GET.get('is_active', '')
        
        # Проверяем, можно ли создавать новые SLA (только для активных вакансий)
        from apps.company_settings.utils import get_active_grades_queryset
        active_vacancies = Vacancy.objects.filter(is_active=True)
        all_grades = get_active_grades_queryset()
        total_possible_slas = active_vacancies.count() * all_grades.count()
        existing_slas_for_active = VacancySLA.objects.filter(vacancy__in=active_vacancies).count()
        
        context['can_create_sla'] = existing_slas_for_active < total_possible_slas
        context['sla_coverage'] = {
            'existing': existing_slas_for_active,
            'total': total_possible_slas,
            'percentage': round((existing_slas_for_active / total_possible_slas * 100), 1) if total_possible_slas > 0 else 0
        }
        
        return context


class VacancySLACreateView(LoginRequiredMixin, CreateView):
    """Создание SLA для вакансии"""
    model = VacancySLA
    form_class = VacancySLAForm
    template_name = 'hiring_plan/sla_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:sla_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'SLA для "{self.object}" успешно создан!')
        return response


class VacancySLAUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование SLA для вакансии"""
    model = VacancySLA
    form_class = VacancySLAForm
    template_name = 'hiring_plan/sla_form.html'
    
    def get_success_url(self):
        return reverse('hiring_plan:sla_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'SLA для "{self.object}" успешно обновлен!')
        return response


# Metrics Views
class MetricsDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard с метриками и KPI"""
    template_name = 'hiring_plan/metrics_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем период из параметров запроса
        period_type = self.request.GET.get('period', 'current_quarter')
        
        # Получаем даты периода
        period_start, period_end = MetricsService.get_period_dates(period_type)
        
        # Рассчитываем метрики
        metrics = MetricsService.calculate_recruitment_metrics(
            period_start, period_end
        )
        
        # Мощность команды (также для выбранного периода)
        team_capacity = MetricsService.get_team_capacity_summary(period_start, period_end)
        
        # KPI Cards
        context['kpi_cards'] = {
            'avg_time_to_offer': metrics.avg_time_to_offer,
            'hiring_velocity': metrics.hiring_velocity_weekly,
            'sla_compliance': metrics.sla_compliance_rate,
            'days_behind_schedule': metrics.avg_days_behind_schedule,
        }
        
        # Данные для графиков
        context['metrics'] = metrics
        context['team_capacity'] = team_capacity
        # Варианты периодов
        period_choices = [
            ('current_month', 'Текущий месяц'),
            ('current_quarter', 'Текущий квартал'),
            ('last_month', 'Прошлый месяц'),
            ('last_quarter', 'Прошлый квартал'),
            ('last_6_months', 'Последние 6 месяцев'),
            ('last_year', 'Прошлый год'),
            ('current_year', 'Текущий год'),
            ('all_time', 'Все время'),
            ('custom', 'Выбрать период вручную'),
        ]
        
        # Описание периода
        period_descriptions = {
            'current_month': 'Текущий месяц',
            'current_quarter': 'Текущий квартал',
            'last_month': 'Прошлый месяц',
            'last_quarter': 'Прошлый квартал',
            'last_6_months': 'Последние 6 месяцев',
            'last_year': 'Прошлый год',
            'current_year': 'Текущий год',
            'all_time': 'Все время',
            'custom': 'Выбрать период вручную',
        }
        
        context['period_info'] = {
            'start': period_start,
            'end': period_end,
            'type': period_type,
            'description': period_descriptions.get(period_type, 'Текущий квартал'),
            'full_description': f'{period_descriptions.get(period_type, "Текущий квартал")} ({period_start.strftime("%d.%m.%Y")} - {period_end.strftime("%d.%m.%Y")})'
        }
        context['period_choices'] = period_choices
        context['current_period'] = period_type
        
        # Прогнозы (для следующего периода того же типа)
        from apps.vacancies.models import Vacancy
        vacancies = Vacancy.objects.all()[:5]
        forecasts = []
        
        # Определяем период для прогноза на основе выбранного периода
        forecast_period = 'next_month'
        if period_type in ['current_quarter', 'last_quarter']:
            forecast_period = 'next_quarter'
        elif period_type in ['current_year', 'last_year']:
            forecast_period = 'next_year'
            
        for vacancy in vacancies:
            try:
                forecast = MetricsService.forecast_demand(vacancy, forecast_period=forecast_period)
                forecasts.append(forecast)
            except:
                pass  # Игнорируем ошибки прогнозирования
        context['forecasts'] = forecasts
        
        return context


class MetricsListView(LoginRequiredMixin, ListView):
    """Список всех метрик"""
    model = RecruitmentMetrics
    template_name = 'hiring_plan/metrics_list.html'
    context_object_name = 'metrics_list'
    paginate_by = 20
    
    def get_queryset(self):
        return RecruitmentMetrics.objects.select_related('vacancy', 'grade').order_by('-period_start')


class ForecastsListView(LoginRequiredMixin, ListView):
    """Список прогнозов"""
    model = DemandForecast
    template_name = 'hiring_plan/forecasts_list.html'
    context_object_name = 'forecasts'
    paginate_by = 20
    
    def get_queryset(self):
        return DemandForecast.objects.select_related('vacancy', 'grade', 'created_by').order_by('-forecast_start')


class RecruiterCapacityListView(LoginRequiredMixin, ListView):
    """Список мощностей рекрутеров"""
    model = RecruiterCapacity
    template_name = 'hiring_plan/recruiter_capacity_list.html'
    context_object_name = 'capacities'
    paginate_by = 20
    
    def get_queryset(self):
        return RecruiterCapacity.objects.select_related('recruiter').order_by('-period_start', 'recruiter')


@login_required
def get_available_grades(request):
    """AJAX endpoint для получения доступных грейдов для вакансии"""
    vacancy_id = request.GET.get('vacancy_id')
    
    if not vacancy_id:
        return JsonResponse({'error': 'vacancy_id is required'}, status=400)
    
    try:
        from apps.vacancies.models import Vacancy
        from apps.company_settings.utils import get_active_grades_queryset
        
        # Получаем вакансию
        vacancy = Vacancy.objects.get(id=vacancy_id)
        
        # Получаем активные грейды компании
        all_grades = get_active_grades_queryset()
        
        # Получаем грейды, для которых уже созданы SLA
        existing_grades = VacancySLA.objects.filter(vacancy=vacancy).values_list('grade', flat=True)
        
        # Находим доступные грейды
        available_grades = all_grades.exclude(id__in=existing_grades)
        
        # Формируем ответ
        data = {
            'available_grades': [
                {'id': grade.id, 'name': grade.name}
                for grade in available_grades
            ],
            'total_grades': all_grades.count(),
            'existing_grades': len(existing_grades),
            'available_count': available_grades.count()
        }
        
        return JsonResponse(data)
        
    except Vacancy.DoesNotExist:
        return JsonResponse({'error': 'Vacancy not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


class YearlyHiringPlanView(LoginRequiredMixin, TemplateView):
    """Годовая таблица заявок с цветными ячейками по месяцам"""
    template_name = 'hiring_plan/yearly_hiring_plan.html'
    
    @staticmethod
    def _status_as_of(request, as_of_date):
        """
        Статус заявки на конец указанной даты (as_of_date).

        Правило:
        - если заявка закрыта/отменена и closed_date <= as_of_date → показываем closed/cancelled
        - иначе если opening_date > as_of_date → planned
        - иначе → in_progress
        """
        # Закрыта/отменена к указанной дате
        if getattr(request, 'closed_date', None) and request.closed_date <= as_of_date:
            return request.status if request.status in ['closed', 'cancelled'] else 'closed'

        # Еще не началась к указанной дате
        if getattr(request, 'opening_date', None) and request.opening_date > as_of_date:
            return 'planned'

        return 'in_progress'

    @staticmethod
    def _effective_period_end(year):
        """Конец периода для расчетов (для текущего года — сегодня, иначе 31.12)."""
        from datetime import datetime
        year_end = datetime(year, 12, 31).date()
        today = timezone.now().date()
        return min(year_end, today) if year == timezone.now().year else year_end

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем год из параметров (по умолчанию текущий год)
        year = int(self.request.GET.get('year', timezone.now().year))
        context['year'] = year
        
        # Период выбранного года
        from datetime import datetime
        period_start = datetime(year, 1, 1).date()
        period_end = datetime(year, 12, 31).date()
        effective_period_end = self._effective_period_end(year)

        # Получаем все заявки, которые пересекают период года:
        # opening_date <= конец года AND (closed_date IS NULL OR closed_date >= начало года)
        requests = HiringRequest.objects.filter(
            opening_date__lte=period_end
        ).filter(
            models.Q(closed_date__isnull=True) | models.Q(closed_date__gte=period_start)
        ).select_related('vacancy', 'grade', 'sla', 'closed_by', 'recruiter').order_by('vacancy__name', 'grade__name')
        
        table_data = self._build_table_data(requests, year)
        
        context['table_data'] = table_data
        context['months'] = [
            'Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
            'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек'
        ]
        
        # Доступные годы для фильтра
        # Получаем годы из существующих заявок
        years_from_db = list(HiringRequest.objects.values_list('opening_date__year', flat=True).distinct())
        
        # Добавляем текущий год и следующий год, если их еще нет в списке
        current_year = timezone.now().year
        next_year = current_year + 1
        
        years_set = set(years_from_db)
        if current_year not in years_set:
            years_set.add(current_year)
        if next_year not in years_set:
            years_set.add(next_year)
        
        # Сортируем по убыванию
        years = sorted(years_set, reverse=True)
        context['available_years'] = years
        
        # Рассчитываем медианы
        context['medians'] = self._calculate_medians(requests, year)
        
        return context

    def _build_table_data(self, requests, year):
        """Собрать строки таблицы годового плана по тем же правилам, что и UI."""
        from datetime import datetime
        effective_period_end = self._effective_period_end(year)

        table_data = []
        for request in requests:
            status_as_of_year_end = self._status_as_of(request, effective_period_end)

            # Считаем общее количество дней работы (накапливается, включая предыдущие годы)
            days_in_year = 0
            if status_as_of_year_end in ['in_progress', 'closed', 'cancelled']:
                start_date = request.opening_date
                if request.closed_date and request.closed_date <= effective_period_end:
                    end_date = request.closed_date
                else:
                    end_date = effective_period_end

                days_in_year = (end_date - start_date).days + 1
                if days_in_year < 0:
                    days_in_year = 0

            # Time-to-hire на конец периода (накапливается)
            time2hire_as_of_year_end = None
            if status_as_of_year_end in ['in_progress', 'closed', 'cancelled'] and request.opening_date:
                if request.closed_date and request.closed_date <= effective_period_end:
                    end_date = request.closed_date
                else:
                    end_date = effective_period_end
                time2hire_as_of_year_end = (end_date - request.opening_date).days

            row_data = {
                'request': request,
                'request_id': request.id,
                'vacancy': request.vacancy.name,
                'grade': request.grade.name,
                'project': request.project or '-',
                'sla_days': request.sla.time_to_offer if request.sla else '-',
                'sla_time2hire': request.sla.time_to_hire if request.sla else '-',
                'opening_date': request.opening_date,
                'deadline': request.deadline,
                'status': status_as_of_year_end,
                'days_in_year': days_in_year,
                'months': self._get_monthly_data(request, year),
                'closed_by': request.closed_by if (request.closed_date and request.closed_date <= effective_period_end) else None,
                'recruiter': request.recruiter,
                'time2hire': time2hire_as_of_year_end
            }
            table_data.append(row_data)
        return table_data
    
    def _calculate_medians(self, requests, year):
        """Рассчитать медианы согласно алгоритму"""
        import statistics
        from datetime import datetime, timedelta
        
        medians = {}
        period_end = datetime(year, 12, 31).date()
        effective_period_end = self._effective_period_end(year)
        
        # 1. Медианный грейд специалистов (по id грейда)
        grade_ids = [request.grade.id for request in requests]
        if grade_ids:
            grade_ids.sort()
            if len(grade_ids) % 2 == 1:
                median_grade_id = grade_ids[len(grade_ids) // 2]
            else:
                mid1 = grade_ids[len(grade_ids) // 2 - 1]
                mid2 = grade_ids[len(grade_ids) // 2]
                median_grade_id = (mid1 + mid2) / 2
            
            # Находим название грейда по ID
            try:
                median_grade = next((request.grade.name for request in requests if request.grade.id == int(median_grade_id)), 
                                  f"ID: {median_grade_id}")
            except ValueError:
                median_grade = f"ID: {median_grade_id}"
            
            medians['grade'] = median_grade
        else:
            medians['grade'] = "—"
        
        # 2. Медиана дней в месяце для вакансий (все переходящие вакансии дробим)
        # Алгоритм: Для каждой вакансии считаем полный срок работы к моменту каждого месяца
        monthly_days = {i: [] for i in range(1, 13)}  # 1-12 месяцы
        
        for request in requests:
            original_start_date = request.opening_date
            end_date = request.closed_date or effective_period_end
            if end_date > effective_period_end:
                end_date = effective_period_end
            
            # Для каждого месяца в выбранном году считаем полный срок работы с заявкой
            for month in range(1, 13):
                # Определяем конец месяца для расчета
                if month == 12:
                    month_end = datetime(year, 12, 31).date()
                else:
                    month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
                
                # Если заявка еще не началась к концу этого месяца, пропускаем
                if original_start_date > month_end:
                    continue
                
                # Если заявка уже закрыта к началу этого месяца, пропускаем
                if request.closed_date and request.closed_date < datetime(year, month, 1).date():
                    continue
                
                # Определяем дату окончания для расчета (конец месяца или дата закрытия)
                calculation_end = min(month_end, end_date)
                
                # Считаем полный срок работы с заявкой к концу этого месяца
                total_days = (calculation_end - original_start_date).days + 1
                
                if total_days > 0:
                    monthly_days[month].append(total_days)
        
        # Рассчитываем медиану для каждого месяца
        monthly_medians = {}
        for month, days_list in monthly_days.items():
            if days_list:
                days_list.sort()
                if len(days_list) % 2 == 1:
                    # Нечетное число - берем средний элемент
                    median_days = days_list[len(days_list) // 2]
                else:
                    # Четное число - среднее из двух средних
                    mid1 = days_list[len(days_list) // 2 - 1]
                    mid2 = days_list[len(days_list) // 2]
                    median_days = (mid1 + mid2) / 2
                monthly_medians[month] = median_days
            else:
                monthly_medians[month] = 0
        
        medians['monthly_days'] = monthly_medians
        
        # 3. Медиана SLA (все значения в днях)
        sla_days = [request.sla.time_to_offer for request in requests if request.sla]
        if sla_days:
            sla_days.sort()
            if len(sla_days) % 2 == 1:
                median_sla = sla_days[len(sla_days) // 2]
            else:
                mid1 = sla_days[len(sla_days) // 2 - 1]
                mid2 = sla_days[len(sla_days) // 2]
                median_sla = (mid1 + mid2) / 2
            medians['sla'] = f"{median_sla:.0f}д"
        else:
            medians['sla'] = "—"
        
        # 4. Медиана "в работе (дней)" - полное количество дней работы с заявкой
        work_days = []
        for request in requests:
            status_as_of = self._status_as_of(request, effective_period_end)
            if status_as_of in ['in_progress', 'closed', 'cancelled']:
                start_date = request.opening_date
                if request.closed_date and request.closed_date <= effective_period_end:
                    end_date = request.closed_date
                else:
                    end_date = effective_period_end

                total_days = (end_date - start_date).days + 1
                if total_days > 0:
                    work_days.append(total_days)
        
        if work_days:
            work_days.sort()
            if len(work_days) % 2 == 1:
                median_work_days = work_days[len(work_days) // 2]
            else:
                mid1 = work_days[len(work_days) // 2 - 1]
                mid2 = work_days[len(work_days) // 2]
                median_work_days = (mid1 + mid2) / 2
            medians['work_days'] = f"{median_work_days:.0f}д"
        else:
            medians['work_days'] = "—"
        
        # 5. Средние значения (аналогично медианам)
        averages = {}
        
        # 5.1. Средний грейд специалистов (по id грейда)
        if grade_ids:
            average_grade_id = sum(grade_ids) / len(grade_ids)
            try:
                average_grade = next((request.grade.name for request in requests if request.grade.id == int(average_grade_id)), 
                                  f"ID: {average_grade_id:.1f}")
            except ValueError:
                average_grade = f"ID: {average_grade_id:.1f}"
            averages['grade'] = average_grade
        else:
            averages['grade'] = "—"
        
        # 5.2. Среднее дней в месяце для вакансий (используем те же данные что и для медианы)
        monthly_averages = {}
        for month, days_list in monthly_days.items():
            if days_list:
                average_days = sum(days_list) / len(days_list)
                monthly_averages[month] = average_days
            else:
                monthly_averages[month] = 0
        averages['monthly_days'] = monthly_averages
        
        # 5.3. Среднее SLA (все значения в днях)
        if sla_days:
            average_sla = sum(sla_days) / len(sla_days)
            averages['sla'] = f"{average_sla:.0f}д"
        else:
            averages['sla'] = "—"
        
        # 5.4. Среднее "в работе (дней)" - общее количество дней работы
        if work_days:
            average_work_days = sum(work_days) / len(work_days)
            averages['work_days'] = f"{average_work_days:.0f}д"
        else:
            averages['work_days'] = "—"
        
        # 5.5. Среднее для T2H
        time2hire_values = []
        for request in requests:
            status_as_of = self._status_as_of(request, effective_period_end)
            if status_as_of in ['in_progress', 'closed', 'cancelled'] and request.opening_date:
                if request.closed_date and request.closed_date <= effective_period_end:
                    end_date = request.closed_date
                else:
                    end_date = effective_period_end
                time2hire_values.append((end_date - request.opening_date).days)
        if time2hire_values:
            average_time2hire = sum(time2hire_values) / len(time2hire_values)
            averages['time2hire'] = f"{average_time2hire:.0f}д"
        else:
            averages['time2hire'] = "—"
        
        # 5.6. Среднее для SLA T2H
        sla_time2hire_values = [request.sla.time_to_hire for request in requests if request.sla and request.sla.time_to_hire]
        if sla_time2hire_values:
            average_sla_time2hire = sum(sla_time2hire_values) / len(sla_time2hire_values)
            averages['sla_time2hire'] = f"{average_sla_time2hire:.0f}д"
        else:
            averages['sla_time2hire'] = "—"
        
        # 6. Медиана для T2H
        if time2hire_values:
            time2hire_values.sort()
            if len(time2hire_values) % 2 == 1:
                median_time2hire = time2hire_values[len(time2hire_values) // 2]
            else:
                mid1 = time2hire_values[len(time2hire_values) // 2 - 1]
                mid2 = time2hire_values[len(time2hire_values) // 2]
                median_time2hire = (mid1 + mid2) / 2
            medians['time2hire'] = f"{median_time2hire:.0f}д"
        else:
            medians['time2hire'] = "—"
        
        # 7. Медиана для SLA T2H
        if sla_time2hire_values:
            sla_time2hire_values.sort()
            if len(sla_time2hire_values) % 2 == 1:
                median_sla_time2hire = sla_time2hire_values[len(sla_time2hire_values) // 2]
            else:
                mid1 = sla_time2hire_values[len(sla_time2hire_values) // 2 - 1]
                mid2 = sla_time2hire_values[len(sla_time2hire_values) // 2]
                median_sla_time2hire = (mid1 + mid2) / 2
            medians['sla_time2hire'] = f"{median_sla_time2hire:.0f}д"
        else:
            medians['sla_time2hire'] = "—"
        
        # 8. Процент закрытых вакансий
        total_requests = len(requests)
        closed_requests = len([r for r in requests if self._status_as_of(r, effective_period_end) == 'closed'])
        if total_requests > 0:
            closed_percentage = (closed_requests / total_requests) * 100
            medians['closed_percentage'] = f"{closed_percentage:.1f}%"
        else:
            medians['closed_percentage'] = "0%"
        
        # Добавляем средние значения в результат
        medians['averages'] = averages
        
        return medians
    
    def _get_monthly_data(self, request, year):
        """Получить данные по месяцам для заявки"""
        months = {}
        
        # Определяем период работы над заявкой
        start_date = request.opening_date
        end_date = request.closed_date or timezone.now().date()

        # Для текущего года не показываем "будущее" сверх сегодняшнего дня
        if year == timezone.now().year:
            today = timezone.now().date()
            if end_date > today:
                end_date = today
        
        # Для планируемых заявок показываем только месяц планируемого открытия
        if request.status == 'planned':
            planned_month = start_date.month
            if start_date.year == year:
                for month in range(1, 13):
                    if month == planned_month:
                        months[month] = {
                            'color': 'lightblue',
                            'active': True,
                            'days': 0  # Планируемые заявки еще не начались
                        }
                    else:
                        months[month] = {
                            'color': 'transparent',
                            'active': False,
                            'days': 0
                        }
            else:
                # Если планируемая заявка не в этом году, не показываем
                for month in range(1, 13):
                    months[month] = {
                        'color': 'transparent',
                        'active': False,
                        'days': 0
                    }
            return months
        
        # Если заявка открыта в другом году, начинаем с января (переходящие заявки)
        if start_date.year < year:
            start_date = timezone.datetime(year, 1, 1).date()
        
        # Если заявка закрыта в другом году, заканчиваем в декабре
        if end_date.year > year:
            end_date = timezone.datetime(year, 12, 31).date()
        
        # Заполняем месяцы
        for month in range(1, 13):
            month_start = timezone.datetime(year, month, 1).date()
            if month == 12:
                month_end = timezone.datetime(year, 12, 31).date()
            else:
                month_end = timezone.datetime(year, month + 1, 1).date() - timezone.timedelta(days=1)
            
            # Проверяем, пересекается ли заявка с этим месяцем
            if start_date <= month_end and end_date >= month_start:
                # Определяем цвет ячейки
                as_of_date = min(month_end, self._effective_period_end(year))
                color = self._get_cell_color(request, as_of_date)
                
                # Вычисляем количество дней в этом месяце
                # Начало периода - максимум из начала заявки и начала месяца
                period_start = max(start_date, month_start)
                # Конец периода - минимум из конца заявки и конца месяца
                period_end = min(end_date, month_end)
                
                # Считаем количество дней
                days_in_month = (period_end - period_start).days + 1
                if days_in_month < 0:
                    days_in_month = 0
                
                months[month] = {
                    'color': color,
                    'active': True,
                    'days': days_in_month
                }
            else:
                months[month] = {
                    'color': 'transparent',
                    'active': False,
                    'days': 0
                }
        
        return months
    
    def _get_cell_color(self, request, as_of_date):
        """Определить цвет ячейки на основе статуса заявки на дату as_of_date"""
        status = self._status_as_of(request, as_of_date)

        if status == 'cancelled':
            return 'gray'
        elif status == 'closed':
            if request.closed_date and request.deadline:
                if request.closed_date <= request.deadline:
                    return 'green'  # Закрыто в срок
                else:
                    return 'red'    # Закрыто с просрочкой
            else:
                return 'green'  # Закрыто (нет дедлайна)
        elif status == 'in_progress':
            # Проверяем, просрочена ли заявка
            if request.deadline and as_of_date > request.deadline:
                return 'yellow'  # Просрочена в работе
            else:
                return 'blue'  # В работе
        elif status == 'planned':
            return 'lightblue'  # Планируется
        else:
            return 'lightgray'  # Остальные статусы


@login_required
def yearly_hiring_plan_export_excel(request):
    """Скачать годовой план найма в Excel (XLSX)."""
    from datetime import datetime
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    year = int(request.GET.get('year', timezone.now().year))
    view = YearlyHiringPlanView()

    period_start = datetime(year, 1, 1).date()
    period_end = datetime(year, 12, 31).date()

    qs = HiringRequest.objects.filter(
        opening_date__lte=period_end
    ).filter(
        models.Q(closed_date__isnull=True) | models.Q(closed_date__gte=period_start)
    ).select_related('vacancy', 'grade', 'sla', 'closed_by', 'recruiter').order_by('vacancy__name', 'grade__name')

    table_data = view._build_table_data(qs, year)
    medians = view._calculate_medians(qs, year)

    months_short = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']

    # Цвета как в UI
    fills = {
        'green': PatternFill('solid', fgColor='D4EDDA'),
        'red': PatternFill('solid', fgColor='F8D7DA'),
        'blue': PatternFill('solid', fgColor='CCE5FF'),
        'lightblue': PatternFill('solid', fgColor='E3F2FD'),
        'yellow': PatternFill('solid', fgColor='FFF3CD'),
        'gray': PatternFill('solid', fgColor='E9ECEF'),
        'lightgray': PatternFill('solid', fgColor='F8F9FA'),
        'transparent': None,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = f"План {year}"

    headers = ['Вакансия', 'Грейд', 'Проект', *months_short, 'SLA (д)', 'Факт (д)', 'T2H | SLA', 'Рекрутер', 'Статус']
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='F8F9FA')
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _status_label(s):
        return {
            'planned': 'Планируется',
            'in_progress': 'В работе',
            'closed': 'Закрыто',
            'cancelled': 'Отменено',
        }.get(s, str(s))

    # Data rows
    for row in table_data:
        t2h_part = '—'
        if row['status'] != 'planned':
            t2h = row.get('time2hire')
            sla_t2h = row.get('sla_time2hire')
            parts = []
            parts.append(f"T2H: {t2h}д" if t2h is not None else "T2H: —")
            parts.append(f"SLA: {sla_t2h}д" if sla_t2h != '-' else "SLA: —")
            t2h_part = '; '.join(parts)

        values = [
            row['vacancy'],
            row['grade'],
            row['project'],
        ]
        # Months
        for m in range(1, 13):
            md = row['months'].get(m, {'active': False, 'days': 0, 'color': 'transparent'})
            values.append(md.get('days') if md.get('active') and md.get('days') else '')
        values += [
            row['sla_days'] if row['sla_days'] != '-' else '',
            row['days_in_year'] if row['status'] != 'planned' else '',
            t2h_part,
            (row['recruiter'].get_full_name() if row.get('recruiter') else '') if row.get('recruiter') else '',
            _status_label(row['status']),
        ]

        ws.append(values)
        current_row = ws.max_row

        # Apply month fills
        month_start_col = 4
        for m in range(1, 13):
            md = row['months'].get(m, {'color': 'transparent'})
            fill = fills.get(md.get('color', 'transparent'))
            if fill:
                ws.cell(row=current_row, column=month_start_col + (m - 1)).fill = fill

    # Medians row
    ws.append([])
    ws.append(['Медианы', medians.get('grade', ''), '', *[
        (medians.get('monthly_days', {}).get(m, 0) or 0) for m in range(1, 13)
    ], medians.get('sla', ''), medians.get('work_days', ''), f"T2H: {medians.get('time2hire','')}; SLA: {medians.get('sla_time2hire','')}", '', medians.get('closed_percentage', '')])

    # Basic formatting
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Column widths
    widths = {
        1: 40,  # Вакансия
        2: 10,  # Грейд
        3: 18,  # Проект
        16: 10, # SLA
        17: 10, # Факт
        18: 28, # T2H | SLA
        19: 18, # Рекрутер
        20: 14, # Статус
    }
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 7)

    # Align months center
    for r in range(2, ws.max_row + 1):
        for c in range(4, 16):  # month columns
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"hiring_plan_{year}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return resp


@login_required
@require_http_methods(['GET'])
def hiring_requests_export_json(request):
    """Скачивание заявок и SLA в виде JSON-файла."""
    from .export_import import export_hiring_requests_json
    data = export_hiring_requests_json()
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    filename = f'hiring_requests_export_{timezone.now().strftime("%Y%m%d_%H%M")}.json'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(['POST'])
def hiring_requests_import_json(request):
    """Импорт заявок и SLA из загруженного JSON-файла."""
    from .export_import import import_hiring_requests_json

    if not request.FILES.get('json_file'):
        messages.error(request, 'Выберите JSON-файл для импорта.')
        return redirect('hiring_plan:hiring_requests_list')

    try:
        raw = request.FILES['json_file'].read().decode('utf-8')
        data = json.loads(raw)
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        messages.error(request, f'Неверный формат JSON: {e}')
        return redirect('hiring_plan:hiring_requests_list')

    sla_created, sla_updated, req_created, errors = import_hiring_requests_json(data)
    if errors:
        for err in errors[:10]:
            messages.warning(request, err)
        if len(errors) > 10:
            messages.warning(request, f'... и ещё {len(errors) - 10} ошибок.')
    if sla_created or sla_updated or req_created:
        messages.success(
            request,
            f'Импорт завершён: SLA создано {sla_created}, обновлено {sla_updated}; заявок создано {req_created}.'
        )
    elif not errors:
        messages.info(request, 'Нет данных для импорта.')
    return redirect('hiring_plan:hiring_requests_list')


