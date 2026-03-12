"""
Модуль для экспорта отчетов в Excel с таблицами и графиками
"""
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class ExcelReportExporter:
    """Класс для экспорта отчетов в Excel"""
    
    def __init__(self, report_data: Dict, report_type: str, report_title: str = None):
        """
        Инициализация экспортера
        
        Args:
            report_data: Данные отчета
            report_type: Тип отчета ('company', 'recruiter', 'vacancy', 'interviewer', 'recruiters_summary')
            report_title: Заголовок отчета
        """
        self.report_data = report_data
        self.report_type = report_type
        self.report_title = report_title or self._get_default_title()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Отчет"
        
    def _get_default_title(self) -> str:
        """Получить заголовок по умолчанию"""
        titles = {
            'company': 'Отчет по компании',
            'recruiter': 'Отчет по рекрутеру',
            'vacancy': 'Отчет по вакансии',
            'interviewer': 'Отчет по интервьюеру',
            'recruiters_summary': 'Сводный отчет по рекрутерам',
        }
        return titles.get(self.report_type, 'Отчет')
    
    def _format_duration_minutes(self, minutes: int) -> str:
        """Форматировать минуты в читаемый формат"""
        if not minutes:
            return "0 ч 0 мин"
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours} ч {mins} мин"
    
    def _apply_header_style(self, cell):
        """Применить стиль заголовка"""
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_cell_style(self, cell, is_even=False):
        """Применить стиль ячейки"""
        if is_even:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def export(self) -> HttpResponse:
        """Экспортировать отчет в Excel"""
        # Заголовок отчета
        self._add_title()
        
        # Общая статистика
        self._add_summary()
        
        # Для отчета по компании - детальная таблица и графики
        if self.report_type == 'company':
            self._add_detailed_company_table()
            self._add_company_charts()
        else:
            # Для других отчетов - стандартная таблица
            self._add_periods_table()
            self._add_chart()
        
        # Дополнительные данные в зависимости от типа отчета
        if self.report_type == 'recruiter':
            self._add_vacancy_stats()
        elif self.report_type == 'recruiters_summary':
            self._add_recruiters_table()
        
        # Настройка ширины столбцов
        self._adjust_column_widths()
        
        # Создаем HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{self.report_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Сохраняем в BytesIO
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        response.write(output.read())
        
        return response
    
    def _add_title(self):
        """Добавить заголовок отчета"""
        self.ws.merge_cells('A1:E1')
        title_cell = self.ws['A1']
        title_cell.value = self.report_title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Период отчета
        start_date = self.report_data.get('start_date')
        end_date = self.report_data.get('end_date')
        if start_date and end_date:
            # Преобразуем date/datetime в строку
            from datetime import date as date_type
            if isinstance(start_date, date_type):
                start_date_str = start_date.strftime('%d.%m.%Y')
            elif isinstance(start_date, datetime):
                start_date_str = start_date.strftime('%d.%m.%Y')
            else:
                start_date_str = str(start_date)
            
            if isinstance(end_date, date_type):
                end_date_str = end_date.strftime('%d.%m.%Y')
            elif isinstance(end_date, datetime):
                end_date_str = end_date.strftime('%d.%m.%Y')
            else:
                end_date_str = str(end_date)
            
            period_str = f"Период: {start_date_str} - {end_date_str}"
            self.ws.merge_cells('A2:E2')
            period_cell = self.ws['A2']
            period_cell.value = period_str
            period_cell.font = Font(size=12)
            period_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.current_row = 4
        else:
            self.current_row = 3
    
    def _add_summary(self):
        """Добавить общую статистику"""
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:E{row}')
        summary_title = self.ws[f'A{row}']
        summary_title.value = "Общая статистика"
        summary_title.font = Font(bold=True, size=14)
        summary_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Данные статистики
        stats = [
            ('Всего скринингов', self.report_data.get('total_screenings', 0)),
            ('Всего интервью', self.report_data.get('total_interviews', 0)),
        ]
        
        if 'total_events' in self.report_data:
            stats.insert(0, ('Всего событий', self.report_data.get('total_events', 0)))
        
        total_time = self.report_data.get('total_time_minutes', 0)
        stats.append(('Суммарное время', self._format_duration_minutes(total_time)))
        
        for i, (label, value) in enumerate(stats):
            col = 'A' if i < 2 else 'D'
            cell_row = row + (i % 2)
            
            label_cell = self.ws[f'{col}{cell_row}']
            label_cell.value = label
            label_cell.font = Font(bold=True)
            self._apply_cell_style(label_cell, is_even=(i % 2 == 1))
            
            value_cell = self.ws[f'{chr(ord(col[0]) + 1)}{cell_row}']
            value_cell.value = value
            self._apply_cell_style(value_cell, is_even=(i % 2 == 1))
        
        self.current_row = row + 3
    
    def _add_periods_table(self):
        """Добавить таблицу данных по периодам"""
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:E{row}')
        table_title = self.ws[f'A{row}']
        table_title.value = "Детализация по периодам"
        table_title.font = Font(bold=True, size=14)
        table_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Заголовки таблицы
        headers = ['Период', 'Скрининги', 'Интервью', 'Всего', 'Время']
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные таблицы
        grouped_data = self.report_data.get('grouped_data', {})
        sorted_periods = sorted(grouped_data.keys())
        
        for period_key in sorted_periods:
            period_data = grouped_data[period_key]
            screenings = period_data.get('screenings', 0)
            interviews = period_data.get('interviews', 0)
            total = screenings + interviews
            time_minutes = period_data.get('total_time_minutes', 0)
            
            data_row = [
                period_key,
                screenings,
                interviews,
                total,
                self._format_duration_minutes(time_minutes)
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                is_even = (row - self.current_row - 1) % 2 == 1
                self._apply_cell_style(cell, is_even=is_even)
            
            row += 1
        
        if not sorted_periods:
            # Нет данных
            self.ws.merge_cells(f'A{row}:E{row}')
            no_data_cell = self.ws[f'A{row}']
            no_data_cell.value = "Нет данных за выбранный период"
            no_data_cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        
        self.current_row = row + 2
    
    def _add_chart(self):
        """Добавить график"""
        grouped_data = self.report_data.get('grouped_data', {})
        if not grouped_data:
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:E{row}')
        chart_title = self.ws[f'A{row}']
        chart_title.value = "График динамики"
        chart_title.font = Font(bold=True, size=14)
        chart_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Создаем данные для графика
        sorted_periods = sorted(grouped_data.keys())
        
        # Заголовки для данных графика
        chart_headers = ['Период', 'Скрининги', 'Интервью', 'Всего']
        for col_idx, header in enumerate(chart_headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные для графика
        chart_start_row = row
        for period_key in sorted_periods:
            period_data = grouped_data[period_key]
            screenings = period_data.get('screenings', 0)
            interviews = period_data.get('interviews', 0)
            total = screenings + interviews
            
            data_row = [period_key, screenings, interviews, total]
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                self._apply_cell_style(cell)
            row += 1
        
        chart_end_row = row - 1
        
        # Создаем график
        chart = LineChart()
        chart.title = "Динамика скринингов и интервью"
        chart.style = 10
        chart.y_axis.title = 'Количество'
        chart.x_axis.title = 'Период'
        
        # Данные для графика
        data = Reference(self.ws, min_col=2, min_row=chart_start_row-1, max_col=4, max_row=chart_end_row)
        chart.add_data(data, titles_from_data=True)
        
        # Категории (периоды)
        cats = Reference(self.ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
        chart.set_categories(cats)
        
        # Размещаем график
        self.ws.add_chart(chart, f'A{row}')
        
        self.current_row = row + 20
    
    def _add_vacancy_stats(self):
        """Добавить статистику по вакансиям (для отчета по рекрутеру)"""
        vacancy_stats = self.report_data.get('vacancy_stats', [])
        if not vacancy_stats:
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:E{row}')
        stats_title = self.ws[f'A{row}']
        stats_title.value = "Статистика по вакансиям"
        stats_title.font = Font(bold=True, size=14)
        stats_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Заголовки таблицы
        headers = ['Вакансия', 'Скрининги', 'Интервью', 'Всего']
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные
        for stat in vacancy_stats:
            vacancy = stat.get('vacancy')
            vacancy_name = vacancy.name if vacancy else 'Не указана'
            screenings = stat.get('screenings', 0)
            interviews = stat.get('interviews', 0)
            total = stat.get('total', 0)
            
            data_row = [vacancy_name, screenings, interviews, total]
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                is_even = (row - self.current_row - 1) % 2 == 1
                self._apply_cell_style(cell, is_even=is_even)
            row += 1
        
        self.current_row = row + 2
    
    def _add_recruiters_table(self):
        """Добавить таблицу по рекрутерам (для сводного отчета)"""
        recruiters_data = self.report_data.get('recruiters', [])
        if not recruiters_data:
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:F{row}')
        recruiters_title = self.ws[f'A{row}']
        recruiters_title.value = "Детализация по рекрутерам"
        recruiters_title.font = Font(bold=True, size=14)
        recruiters_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Заголовки таблицы
        headers = ['Рекрутер', 'Скрининги', 'Интервью', 'Всего', 'Время']
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные
        for recruiter_data in recruiters_data:
            recruiter = recruiter_data.get('recruiter')
            recruiter_name = recruiter.get_full_name() if hasattr(recruiter, 'get_full_name') else recruiter.username
            screenings = recruiter_data.get('screenings', 0)
            interviews = recruiter_data.get('interviews', 0)
            total = recruiter_data.get('total', 0)
            time_minutes = recruiter_data.get('total_time_minutes', 0)
            
            data_row = [
                recruiter_name,
                screenings,
                interviews,
                total,
                self._format_duration_minutes(time_minutes)
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                is_even = (row - self.current_row - 1) % 2 == 1
                self._apply_cell_style(cell, is_even=is_even)
            row += 1
        
        self.current_row = row + 2
    
    def _add_detailed_company_table(self):
        """Добавить детальную таблицу для отчета по компании"""
        detailed_data = self.report_data.get('detailed_data', [])
        if not detailed_data:
            # Если нет детальных данных, используем стандартную таблицу
            self._add_periods_table()
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:G{row}')
        table_title = self.ws[f'A{row}']
        table_title.value = "Детализация по периодам и рекрутерам"
        table_title.font = Font(bold=True, size=14)
        table_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Заголовки таблицы
        headers = ['Период', 'Рекрутер', 'Интервьюеры', 'Скрининги', 'Интервью', 'Сумма', 'Время']
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные таблицы
        for data in detailed_data:
            data_row = [
                data.get('period', ''),
                data.get('recruiter_name', ''),
                data.get('interviewer_names', '—'),
                data.get('screenings', 0),
                data.get('interviews', 0),
                data.get('total', 0),
                self._format_duration_minutes(data.get('total_time_minutes', 0))
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                is_even = (row - self.current_row - 1) % 2 == 1
                self._apply_cell_style(cell, is_even=is_even)
            
            row += 1
        
        self.current_row = row + 2
    
    def _add_company_charts(self):
        """Добавить графики для отчета по компании"""
        detailed_data = self.report_data.get('detailed_data', [])
        grouped_data = self.report_data.get('grouped_data', {})
        period = self.report_data.get('period', 'monthly')
        
        if not detailed_data and not grouped_data:
            return
        
        # График 1: Динамика по периодам (скрининги и интервью)
        self._add_period_chart(grouped_data)
        
        # График 2: По рекрутерам (общая статистика)
        self._add_recruiters_chart(detailed_data)
        
        # График 3: Скрининги vs Интервью по периодам
        self._add_screenings_vs_interviews_chart(grouped_data)
    
    def _add_period_chart(self, grouped_data):
        """Добавить график динамики по периодам"""
        if not grouped_data:
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:G{row}')
        chart_title = self.ws[f'A{row}']
        chart_title.value = "График 1: Динамика скринингов и интервью по периодам"
        chart_title.font = Font(bold=True, size=14)
        chart_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        sorted_periods = sorted(grouped_data.keys())
        
        # Заголовки для данных графика
        chart_headers = ['Период', 'Скрининги', 'Интервью', 'Всего']
        for col_idx, header in enumerate(chart_headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные для графика
        chart_start_row = row
        for period_key in sorted_periods:
            period_data = grouped_data[period_key]
            screenings = period_data.get('screenings', 0)
            interviews = period_data.get('interviews', 0)
            total = screenings + interviews
            
            data_row = [period_key, screenings, interviews, total]
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                self._apply_cell_style(cell)
            row += 1
        
        chart_end_row = row - 1
        
        # Создаем график
        chart = LineChart()
        chart.title = "Динамика скринингов и интервью"
        chart.style = 10
        chart.y_axis.title = 'Количество'
        chart.x_axis.title = 'Период'
        
        # Данные для графика
        data = Reference(self.ws, min_col=2, min_row=chart_start_row-1, max_col=4, max_row=chart_end_row)
        chart.add_data(data, titles_from_data=True)
        
        # Категории (периоды)
        cats = Reference(self.ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
        chart.set_categories(cats)
        
        # Размещаем график
        self.ws.add_chart(chart, f'A{row}')
        
        self.current_row = row + 20
    
    def _add_recruiters_chart(self, detailed_data):
        """Добавить график по рекрутерам"""
        if not detailed_data:
            return
        
        # Группируем данные по рекрутерам
        recruiters_stats = {}
        for data in detailed_data:
            recruiter_name = data.get('recruiter_name', '')
            if recruiter_name not in recruiters_stats:
                recruiters_stats[recruiter_name] = {
                    'screenings': 0,
                    'interviews': 0,
                    'total': 0,
                    'total_time_minutes': 0,
                }
            
            recruiters_stats[recruiter_name]['screenings'] += data.get('screenings', 0)
            recruiters_stats[recruiter_name]['interviews'] += data.get('interviews', 0)
            recruiters_stats[recruiter_name]['total'] += data.get('total', 0)
            recruiters_stats[recruiter_name]['total_time_minutes'] += data.get('total_time_minutes', 0)
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:G{row}')
        chart_title = self.ws[f'A{row}']
        chart_title.value = "График 2: Статистика по рекрутерам"
        chart_title.font = Font(bold=True, size=14)
        chart_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Заголовки для данных графика
        chart_headers = ['Рекрутер', 'Скрининги', 'Интервью', 'Всего']
        for col_idx, header in enumerate(chart_headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные для графика
        chart_start_row = row
        sorted_recruiters = sorted(recruiters_stats.items(), key=lambda x: x[1]['total'], reverse=True)
        
        for recruiter_name, stats in sorted_recruiters:
            data_row = [
                recruiter_name,
                stats['screenings'],
                stats['interviews'],
                stats['total']
            ]
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                self._apply_cell_style(cell)
            row += 1
        
        chart_end_row = row - 1
        
        if chart_end_row >= chart_start_row:
            # Создаем график (столбчатая диаграмма)
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Статистика по рекрутерам"
            chart.y_axis.title = 'Количество'
            chart.x_axis.title = 'Рекрутер'
            
            # Данные для графика
            data = Reference(self.ws, min_col=2, min_row=chart_start_row-1, max_col=4, max_row=chart_end_row)
            chart.add_data(data, titles_from_data=True)
            
            # Категории (рекрутеры)
            cats = Reference(self.ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
            chart.set_categories(cats)
            
            # Размещаем график
            self.ws.add_chart(chart, f'A{row}')
        
        self.current_row = row + 20
    
    def _add_screenings_vs_interviews_chart(self, grouped_data):
        """Добавить график сравнения скринингов и интервью"""
        if not grouped_data:
            return
        
        row = self.current_row
        
        # Заголовок секции
        self.ws.merge_cells(f'A{row}:G{row}')
        chart_title = self.ws[f'A{row}']
        chart_title.value = "График 3: Сравнение скринингов и интервью по периодам"
        chart_title.font = Font(bold=True, size=14)
        chart_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        sorted_periods = sorted(grouped_data.keys())
        
        # Заголовки для данных графика
        chart_headers = ['Период', 'Скрининги', 'Интервью']
        for col_idx, header in enumerate(chart_headers, start=1):
            cell = self.ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_header_style(cell)
        row += 1
        
        # Данные для графика
        chart_start_row = row
        for period_key in sorted_periods:
            period_data = grouped_data[period_key]
            screenings = period_data.get('screenings', 0)
            interviews = period_data.get('interviews', 0)
            
            data_row = [period_key, screenings, interviews]
            for col_idx, value in enumerate(data_row, start=1):
                cell = self.ws.cell(row=row, column=col_idx)
                cell.value = value
                self._apply_cell_style(cell)
            row += 1
        
        chart_end_row = row - 1
        
        # Создаем график
        chart = LineChart()
        chart.title = "Скрининги vs Интервью"
        chart.style = 10
        chart.y_axis.title = 'Количество'
        chart.x_axis.title = 'Период'
        
        # Данные для графика
        data = Reference(self.ws, min_col=2, min_row=chart_start_row-1, max_col=3, max_row=chart_end_row)
        chart.add_data(data, titles_from_data=True)
        
        # Категории (периоды)
        cats = Reference(self.ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
        chart.set_categories(cats)
        
        # Размещаем график
        self.ws.add_chart(chart, f'A{row}')
        
        self.current_row = row + 20
    
    def _adjust_column_widths(self):
        """Настроить ширину столбцов"""
        if self.report_type == 'company':
            column_widths = {
                'A': 20,  # Период
                'B': 25,  # Рекрутер
                'C': 40,  # Интервьюеры
                'D': 12,  # Скрининги
                'E': 12,  # Интервью
                'F': 12,  # Сумма
                'G': 15,  # Время
            }
        else:
            column_widths = {
                'A': 25,  # Период/Вакансия/Рекрутер
                'B': 15,  # Скрининги
                'C': 15,  # Интервью
                'D': 15,  # Всего
                'E': 20,  # Время
                'F': 20,  # Дополнительные столбцы
            }
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width

