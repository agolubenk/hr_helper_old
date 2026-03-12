"""
Views для отчетности
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
import json
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

from apps.reporting.services import ReportGenerator
from apps.vacancies.models import Vacancy
from apps.interviewers.models import Interviewer
from apps.reporting.models import CalendarEvent
from apps.google_oauth.models import GoogleOAuthAccount
from apps.google_oauth.services import GoogleOAuthService, GoogleCalendarService
from apps.reporting.excel_export import ExcelReportExporter

User = get_user_model()


def parse_date_range(request, period='monthly'):
    """
    Вспомогательная функция для парсинга диапазона дат из запроса
    
    Returns:
        tuple: (start_date, end_date) - оба datetime с timezone
    """
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Если даты не указаны, используем текущий период
    if not start_date_str or not end_date_str:
        end_date = timezone.now()
        if period == 'daily':
            start_date = end_date - timedelta(days=30)
        elif period == 'weekly':
            start_date = end_date - timedelta(weeks=12)
        elif period == 'monthly':
            start_date = end_date - relativedelta(months=12)
        elif period == 'quarterly':
            start_date = end_date - relativedelta(months=12)
        elif period == 'yearly':
            start_date = end_date - relativedelta(years=5)
        else:
            start_date = end_date - relativedelta(months=12)
    else:
        try:
            # Парсим даты и конвертируем в aware datetime
            start_date_naive = datetime.fromisoformat(start_date_str)
            end_date_naive = datetime.fromisoformat(end_date_str)
            
            # Если datetime naive, добавляем timezone
            if timezone.is_naive(start_date_naive):
                start_date = timezone.make_aware(start_date_naive)
            else:
                start_date = start_date_naive
            
            if timezone.is_naive(end_date_naive):
                end_date = timezone.make_aware(end_date_naive)
            else:
                end_date = end_date_naive
            
            # Устанавливаем время начала дня для start_date и конец дня для end_date
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        except:
            end_date = timezone.now()
            start_date = end_date - relativedelta(months=12)
    
    return start_date, end_date


@login_required
def report_dashboard(request):
    """Главная страница отчетности"""
    return render(request, 'reporting/dashboard.html')


@login_required
def export_calendar_events_json(request):
    """Экспорт событий календаря отчётности в JSON. Существующие при импорте перезаписываются по event_id."""
    from .export_import import export_calendar_events_json as do_export
    data = do_export()
    response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporting_calendar_events.json"'
    return response


@login_required
def import_calendar_events_json(request):
    """Импорт событий календаря из JSON. Существующие записи перезаписываются по event_id."""
    from .export_import import import_calendar_events_json as do_import
    from django.contrib import messages
    if request.method != 'POST':
        messages.error(request, 'Неверный метод запроса.')
        return redirect('reporting:dashboard')
    data = None
    if request.FILES.get('json_file'):
        try:
            raw = request.FILES['json_file'].read().decode('utf-8')
            data = json.loads(raw)
        except (UnicodeDecodeError, json.JSONDecodeError) as e:
            messages.error(request, f'Ошибка чтения JSON: {e}')
            return redirect('reporting:dashboard')
    elif request.POST.get('json_data'):
        try:
            data = json.loads(request.POST['json_data'])
        except json.JSONDecodeError as e:
            messages.error(request, f'Ошибка формата JSON: {e}')
            return redirect('reporting:dashboard')
    if not data:
        messages.error(request, 'Загрузите JSON-файл или вставьте JSON в поле.')
        return redirect('reporting:dashboard')
    created, updated, errors = do_import(data)
    if errors:
        messages.warning(request, f'Импорт завершён с ошибками: создано {created}, обновлено {updated}. Ошибки: {"; ".join(errors[:5])}{"..." if len(errors) > 5 else ""}.')
    else:
        messages.success(request, f'Импорт завершён: создано {created}, обновлено {updated}.')
    return redirect('reporting:dashboard')


@login_required
def company_report(request):
    """Отчет по компании"""
    # Получаем параметры фильтров
    period = request.GET.get('period', 'monthly')  # daily, weekly, monthly, quarterly, yearly
    
    # Парсим диапазон дат
    start_date, end_date = parse_date_range(request, period)
    
    # Генерируем отчет
    generator = ReportGenerator(request.user)
    report_data = generator.generate_company_report(start_date, end_date, period)
    
    # Получаем список рекрутеров для фильтра графика
    recruiters = User.objects.filter(groups__name='Рекрутер').distinct()
    
    # Получаем список интервьюеров для фильтра графика
    from apps.interviewers.models import Interviewer
    interviewers = Interviewer.objects.filter(is_active=True)
    
    # Получаем список вакансий для фильтра графика
    from apps.vacancies.models import Vacancy
    vacancies = Vacancy.objects.all().order_by('name')
    
    context = {
        'report_data': report_data,
        'period': period,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'recruiters': recruiters,
        'interviewers': interviewers,
        'vacancies': vacancies,
    }
    
    return render(request, 'reporting/company_report.html', context)


@login_required
def recruiters_summary_report(request):
    """Сводный отчет по всем рекрутерам с разбивкой по скринингам и интервью"""
    # Получаем параметры фильтров
    period = request.GET.get('period', 'monthly')
    
    # Парсим диапазон дат
    start_date, end_date = parse_date_range(request, period)
    
    # Генерируем отчет
    generator = ReportGenerator(request.user)
    report_data = generator.generate_recruiters_summary_report(start_date, end_date, period)
    
    context = {
        'report_data': report_data,
        'period': period,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
    }
    
    return render(request, 'reporting/recruiters_summary_report.html', context)


@login_required
def recruiter_report(request, recruiter_id=None):
    """Отчет по рекрутеру"""
    # Если ID не указан, показываем список рекрутеров
    if not recruiter_id:
        recruiters = User.objects.filter(groups__name='Рекрутер').distinct()
        return render(request, 'reporting/recruiter_list.html', {'recruiters': recruiters})
    
    recruiter = get_object_or_404(User, id=recruiter_id, groups__name='Рекрутер')
    
    # Получаем параметры фильтров
    period = request.GET.get('period', 'monthly')
    
    # Парсим диапазон дат
    start_date, end_date = parse_date_range(request, period)
    
    # Генерируем отчет
    generator = ReportGenerator(request.user)
    report_data = generator.generate_recruiter_report(recruiter, start_date, end_date, period)
    
    # Получаем список интервьюеров для фильтра графика
    interviewers = Interviewer.objects.filter(is_active=True)
    
    context = {
        'report_data': report_data,
        'recruiter': recruiter,
        'period': period,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'interviewers': interviewers,
    }
    
    return render(request, 'reporting/recruiter_report.html', context)


@login_required
def vacancy_report(request, vacancy_id=None):
    """Отчет по вакансии"""
    # Если ID не указан, показываем список вакансий (активные и неактивные)
    if not vacancy_id:
        vacancies = Vacancy.objects.all().select_related('recruiter').order_by('-is_active', '-created_at')
        return render(request, 'reporting/vacancy_list.html', {'vacancies': vacancies})
    
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    
    # Получаем параметры фильтров
    period = request.GET.get('period', 'monthly')
    
    # Парсим диапазон дат
    start_date, end_date = parse_date_range(request, period)
    
    # Генерируем отчет
    generator = ReportGenerator(request.user)
    report_data = generator.generate_vacancy_report(vacancy, start_date, end_date, period)
    
    # Получаем список рекрутеров для фильтра графика
    recruiters = User.objects.filter(groups__name='Рекрутер').distinct()
    
    context = {
        'report_data': report_data,
        'vacancy': vacancy,
        'period': period,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'recruiters': recruiters,
    }
    
    return render(request, 'reporting/vacancy_report.html', context)


@login_required
def interviewer_report(request, interviewer_id=None):
    """Отчет по интервьюеру"""
    # Если ID не указан, показываем список интервьюеров со статистикой
    if not interviewer_id:
        interviewers = Interviewer.objects.filter(is_active=True).order_by('last_name', 'first_name')
        
        # Получаем параметры фильтров для общей статистики
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        # Парсим диапазон дат
        if start_date_str and end_date_str:
            try:
                start_date_naive = datetime.fromisoformat(start_date_str)
                end_date_naive = datetime.fromisoformat(end_date_str)
                
                if timezone.is_naive(start_date_naive):
                    start_date = timezone.make_aware(start_date_naive)
                else:
                    start_date = start_date_naive
                
                if timezone.is_naive(end_date_naive):
                    end_date = timezone.make_aware(end_date_naive)
                else:
                    end_date = end_date_naive
                
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            except:
                # Если ошибка парсинга, используем значения по умолчанию
                end_date = timezone.now()
                start_date = end_date - relativedelta(months=12)
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            # Используем значения по умолчанию - последние 12 месяцев
            end_date = timezone.now()
            start_date = end_date - relativedelta(months=12)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Получаем все события за период
        all_events = CalendarEvent.objects.filter(
            start_time__gte=start_date,
            start_time__lte=end_date
        ).select_related('vacancy').order_by('start_time')
        
        # Собираем статистику по каждому интервьюеру
        interviewer_stats = []
        for interviewer in interviewers:
            interviewer_email_lower = interviewer.email.lower()
            
            # Фильтруем события, где интервьюер является участником
            # ИСКЛЮЧАЕМ события, где рекрутер (владелец) совпадает с интервьюером
            interviewer_events = []
            for event in all_events:
                attendees = event.attendees or []
                is_participant = False
                
                for attendee in attendees:
                    if isinstance(attendee, dict):
                        attendee_email = attendee.get('email', '').lower()
                        if attendee_email == interviewer_email_lower:
                            is_participant = True
                            break
                    elif isinstance(attendee, str):
                        if attendee.lower() == interviewer_email_lower:
                            is_participant = True
                            break
                
                # Исключаем события, где рекрутер также является интервьюером
                if is_participant:
                    # Проверяем, не является ли рекрутер также интервьюером
                    recruiter_is_interviewer = False
                    if event.recruiter and event.recruiter.email:
                        recruiter_email_lower = event.recruiter.email.lower()
                        for attendee in attendees:
                            if isinstance(attendee, dict):
                                attendee_email = attendee.get('email', '').lower()
                            elif isinstance(attendee, str):
                                attendee_email = attendee.lower()
                            else:
                                continue
                            if attendee_email == recruiter_email_lower:
                                recruiter_is_interviewer = True
                                break
                    
                    # Добавляем событие только если рекрутер не является интервьюером
                    if not recruiter_is_interviewer:
                        interviewer_events.append(event)
            
            # Подсчитываем статистику
            screenings = sum(1 for e in interviewer_events if e.event_type == 'screening')
            interviews = sum(1 for e in interviewer_events if e.event_type == 'interview')
            total_time_minutes = sum(e.duration_minutes or 0 for e in interviewer_events)
            
            # Вычисляем конверсию из скринингов в интервью
            conversion_rate = None
            if screenings > 0:
                conversion_rate = round((interviews / screenings) * 100, 2)
            
            interviewer_stats.append({
                'interviewer': interviewer,
                'screenings': screenings,
                'interviews': interviews,
                'total_time_minutes': total_time_minutes,
                'conversion_rate': conversion_rate,
            })
        
        # Получаем параметр сортировки
        sort_by = request.GET.get('sort', 'total_desc')
        
        # Применяем сортировку
        if sort_by == 'name_asc':
            interviewer_stats.sort(key=lambda x: x['interviewer'].get_full_name().lower())
        elif sort_by == 'name_desc':
            interviewer_stats.sort(key=lambda x: x['interviewer'].get_full_name().lower(), reverse=True)
        elif sort_by == 'screenings_asc':
            interviewer_stats.sort(key=lambda x: x['screenings'])
        elif sort_by == 'screenings_desc':
            interviewer_stats.sort(key=lambda x: x['screenings'], reverse=True)
        elif sort_by == 'interviews_asc':
            interviewer_stats.sort(key=lambda x: x['interviews'])
        elif sort_by == 'interviews_desc':
            interviewer_stats.sort(key=lambda x: x['interviews'], reverse=True)
        elif sort_by == 'time_asc':
            interviewer_stats.sort(key=lambda x: x['total_time_minutes'])
        elif sort_by == 'time_desc':
            interviewer_stats.sort(key=lambda x: x['total_time_minutes'], reverse=True)
        elif sort_by == 'conversion_asc':
            interviewer_stats.sort(key=lambda x: x['conversion_rate'] if x['conversion_rate'] is not None else -1)
        elif sort_by == 'conversion_desc':
            interviewer_stats.sort(key=lambda x: x['conversion_rate'] if x['conversion_rate'] is not None else -1, reverse=True)
        else:
            # По умолчанию сортируем по общему количеству встреч (скрининги + интервью)
            interviewer_stats.sort(key=lambda x: x['screenings'] + x['interviews'], reverse=True)
        
        context = {
            'interviewers': interviewers,
            'interviewer_stats': interviewer_stats,
            'start_date': start_date.date(),
            'end_date': end_date.date(),
            'sort_by': sort_by,
        }
        return render(request, 'reporting/interviewer_list.html', context)
    
    interviewer = get_object_or_404(Interviewer, id=interviewer_id)
    
    # Получаем параметры фильтров
    period = request.GET.get('period', 'monthly')
    
    # Парсим диапазон дат
    start_date, end_date = parse_date_range(request, period)
    
    # Генерируем отчет
    generator = ReportGenerator(request.user)
    report_data = generator.generate_interviewer_report(interviewer, start_date, end_date, period)
    
    # Получаем список рекрутеров для фильтра графика
    recruiters = User.objects.filter(groups__name='Рекрутер').distinct()
    
    context = {
        'report_data': report_data,
        'interviewer': interviewer,
        'period': period,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'recruiters': recruiters,
    }
    
    return render(request, 'reporting/interviewer_report.html', context)


@login_required
def api_report_data(request):
    """API endpoint для получения данных отчета в JSON формате"""
    report_type = request.GET.get('report_type') or request.GET.get('type')  # company, recruiter, vacancy, interviewer
    period = request.GET.get('period', 'monthly')
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if not start_date_str or not end_date_str:
        return JsonResponse({'error': 'start_date and end_date are required'}, status=400)
    
    try:
        # Парсим даты с правильной обработкой timezone
        start_date_naive = datetime.fromisoformat(start_date_str)
        end_date_naive = datetime.fromisoformat(end_date_str)
        
        # Если datetime naive, добавляем timezone
        if timezone.is_naive(start_date_naive):
            start_date = timezone.make_aware(start_date_naive)
        else:
            start_date = start_date_naive
        
        if timezone.is_naive(end_date_naive):
            end_date = timezone.make_aware(end_date_naive)
        else:
            end_date = end_date_naive
        
        # Устанавливаем время начала дня для start_date и конец дня для end_date
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    except Exception as e:
        return JsonResponse({'error': f'Invalid date format: {str(e)}'}, status=400)
    
    generator = ReportGenerator(request.user)
    
    try:
        if report_type == 'company':
            recruiter_id = request.GET.get('recruiter_id')
            recruiter_id = int(recruiter_id) if recruiter_id else None
            
            interviewer_id = request.GET.get('interviewer_id')
            interviewer_id = int(interviewer_id) if interviewer_id else None
            
            vacancy_id = request.GET.get('vacancy_id')
            vacancy_id = int(vacancy_id) if vacancy_id else None
            
            report_data = generator.generate_company_report(
                start_date, 
                end_date, 
                period, 
                recruiter_id=recruiter_id,
                interviewer_id=interviewer_id,
                vacancy_id=vacancy_id
            )
        elif report_type == 'recruiters_summary':
            report_data = generator.generate_recruiters_summary_report(start_date, end_date, period)
        elif report_type == 'recruiter':
            recruiter_id = request.GET.get('recruiter_id')
            if not recruiter_id:
                return JsonResponse({'error': 'recruiter_id is required'}, status=400)
            recruiter = get_object_or_404(User, id=recruiter_id)
            interviewer_id = request.GET.get('interviewer_id')
            interviewer_id = int(interviewer_id) if interviewer_id else None
            report_data = generator.generate_recruiter_report(recruiter, start_date, end_date, period, interviewer_id=interviewer_id)
        elif report_type == 'vacancy':
            vacancy_id = request.GET.get('vacancy_id')
            if not vacancy_id:
                return JsonResponse({'error': 'vacancy_id is required'}, status=400)
            vacancy = get_object_or_404(Vacancy, id=vacancy_id)
            recruiter_id = request.GET.get('recruiter_id')
            recruiter_id = int(recruiter_id) if recruiter_id else None
            report_data = generator.generate_vacancy_report(vacancy, start_date, end_date, period, recruiter_id=recruiter_id)
        elif report_type == 'interviewer':
            interviewer_id = request.GET.get('interviewer_id')
            if not interviewer_id:
                return JsonResponse({'error': 'interviewer_id is required'}, status=400)
            interviewer = get_object_or_404(Interviewer, id=interviewer_id)
            recruiter_id = request.GET.get('recruiter_id')
            recruiter_id = int(recruiter_id) if recruiter_id else None
            report_data = generator.generate_interviewer_report(interviewer, start_date, end_date, period, recruiter_id=recruiter_id)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
        
        # Преобразуем данные для JSON (убираем объекты моделей)
        grouped_data = report_data.get('grouped_data', {})
        
        # Очищаем grouped_data от объектов моделей и других несериализуемых типов
        cleaned_grouped_data = {}
        for key, value in grouped_data.items():
            if isinstance(value, dict):
                cleaned_grouped_data[key] = {
                    'screenings': value.get('screenings', 0),
                    'interviews': value.get('interviews', 0),
                    'total_time_minutes': value.get('total_time_minutes', 0),
                }
            else:
                cleaned_grouped_data[key] = value
        
        json_data = {
            'period': report_data.get('period'),
            'start_date': report_data.get('start_date').isoformat() if isinstance(report_data.get('start_date'), datetime) else str(report_data.get('start_date')),
            'end_date': report_data.get('end_date').isoformat() if isinstance(report_data.get('end_date'), datetime) else str(report_data.get('end_date')),
            'total_screenings': report_data.get('total_screenings', 0),
            'total_interviews': report_data.get('total_interviews', 0),
            'total_time_minutes': report_data.get('total_time_minutes', 0),
            'grouped_data': cleaned_grouped_data,
        }
        
        # Для сводного отчета по рекрутерам добавляем дополнительную информацию
        if report_type == 'recruiters_summary':
            json_data['recruiters'] = []
            for recruiter_data in report_data.get('recruiters', []):
                recruiter = recruiter_data.get('recruiter')
                json_data['recruiters'].append({
                    'recruiter_id': recruiter.id if recruiter else None,
                    'recruiter_name': recruiter.get_full_name() or recruiter.username if recruiter else None,
                    'recruiter_email': recruiter.email if recruiter else None,
                    'screenings': recruiter_data.get('screenings', 0),
                    'interviews': recruiter_data.get('interviews', 0),
                    'total': recruiter_data.get('total', 0),
                    'vacancy_stats': [
                        {
                            'vacancy_id': vs.get('vacancy').id if vs.get('vacancy') else None,
                            'vacancy_name': vs.get('vacancy').name if vs.get('vacancy') else None,
                            'screenings': vs.get('screenings', 0),
                            'interviews': vs.get('interviews', 0),
                            'total': vs.get('total', 0),
                        }
                        for vs in recruiter_data.get('vacancy_stats', [])
                    ],
                })
            json_data['total_recruiters'] = report_data.get('total_recruiters', 0)
            json_data['total_events'] = report_data.get('total_events', 0)
        
        # Для отчета по рекрутеру добавляем статистику по вакансиям
        elif report_type == 'recruiter':
            recruiter_obj = report_data.get('recruiter')
            json_data['recruiter_id'] = recruiter_obj.id if recruiter_obj else None
            json_data['recruiter_name'] = (recruiter_obj.get_full_name() or recruiter_obj.username) if recruiter_obj else None
            json_data['total_events'] = report_data.get('total_events', 0)
            json_data['vacancy_stats'] = [
                {
                    'vacancy_id': vs.get('vacancy').id if vs.get('vacancy') else None,
                    'vacancy_name': vs.get('vacancy').name if vs.get('vacancy') else None,
                    'screenings': vs.get('screenings', 0),
                    'interviews': vs.get('interviews', 0),
                    'total': vs.get('total', 0),
                }
                for vs in report_data.get('vacancy_stats', [])
            ]
        
        # Для отчета по интервьюеру добавляем информацию об интервьюере
        elif report_type == 'interviewer':
            interviewer_obj = report_data.get('interviewer')
            json_data['interviewer_id'] = interviewer_obj.id if interviewer_obj else None
            json_data['interviewer_name'] = interviewer_obj.get_full_name() if interviewer_obj else None
            json_data['interviewer_email'] = interviewer_obj.email if interviewer_obj else None
            json_data['total_events'] = report_data.get('total_screenings', 0) + report_data.get('total_interviews', 0)
        
        return JsonResponse(json_data)
    
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Ошибка в api_report_data: {str(e)}")
        print(f"Traceback: {error_traceback}")
        return JsonResponse({'error': str(e), 'traceback': error_traceback}, status=500)


@login_required
def sync_calendar_events(request):
    """API endpoint для синхронизации событий календаря"""
    try:
        from django.db.models import Max
        
        # Определяем период синхронизации из параметров запроса или используем значения по умолчанию
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            # Парсим даты из параметров запроса
            try:
                start_date_naive = datetime.fromisoformat(start_date_str)
                end_date_naive = datetime.fromisoformat(end_date_str)
                
                if timezone.is_naive(start_date_naive):
                    start_date = timezone.make_aware(start_date_naive)
                else:
                    start_date = start_date_naive
                
                if timezone.is_naive(end_date_naive):
                    end_date = timezone.make_aware(end_date_naive)
                else:
                    end_date = end_date_naive
                
                # Устанавливаем время начала дня для start_date и конец дня для end_date
                start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                sync_end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                
                # Ограничиваем период 2025 годом
                year_2025_start = datetime(2025, 1, 1, 0, 0, 0)
                if timezone.is_naive(year_2025_start):
                    year_2025_start = timezone.make_aware(year_2025_start)
                year_2025_end = datetime(2025, 12, 31, 23, 59, 59, 999999)
                if timezone.is_naive(year_2025_end):
                    year_2025_end = timezone.make_aware(year_2025_end)
                
                # Ограничиваем даты 2025 годом
                if start_date < year_2025_start:
                    start_date = year_2025_start
                if sync_end_date > year_2025_end:
                    sync_end_date = year_2025_end
            except Exception as e:
                print(f"⚠️ Ошибка парсинга дат из параметров: {e}. Используем значения по умолчанию.")
                # Используем значения по умолчанию при ошибке парсинга - только 2025 год
                start_date = datetime(2025, 1, 1, 0, 0, 0)
                if timezone.is_naive(start_date):
                    start_date = timezone.make_aware(start_date)
                sync_end_date = datetime(2025, 12, 31, 23, 59, 59, 999999)
                if timezone.is_naive(sync_end_date):
                    sync_end_date = timezone.make_aware(sync_end_date)
        else:
            # Используем значения по умолчанию - только 2025 год
            start_date = datetime(2025, 1, 1, 0, 0, 0)
            if timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            
            # Конец: 31.12.2025 - синхронизируем только события за 2025 год
            sync_end_date = datetime(2025, 12, 31, 23, 59, 59, 999999)
            if timezone.is_naive(sync_end_date):
                sync_end_date = timezone.make_aware(sync_end_date)
        
        # Получаем всех рекрутеров (кроме admin)
        recruiters = User.objects.filter(
            groups__name='Рекрутер'
        ).exclude(username='admin').distinct()
        
        # Дополнительные календари для синхронизации (например, календарь компании)
        additional_calendars = [
            {
                'calendar_id': 'andrey.chernomordin@softnetix.io',
                'name': 'Календарь компании (Andrey Chernomordin)',
                'recruiter': None  # Будет определен автоматически или можно указать конкретного рекрутера
            }
        ]
        
        total_synced = 0
        total_errors = 0
        synced_recruiters = []
        skipped_recruiters = []
        
        print(f"🔄 Начинаем синхронизацию для {recruiters.count()} рекрутеров")
        print(f"📅 Период: {start_date.date()} - {sync_end_date.date()} (конец предыдущей недели)")
        
        # Функция для синхронизации календаря
        def sync_calendar_for_user(user, calendar_id, calendar_name='primary'):
            """Синхронизирует календарь для пользователя"""
            try:
                # Проверяем наличие Google OAuth аккаунта
                try:
                    oauth_account = GoogleOAuthAccount.objects.get(user=user)
                except GoogleOAuthAccount.DoesNotExist:
                    return 0, f'Нет Google OAuth аккаунта'
                
                if not oauth_account.is_token_valid():
                    return 0, f'Токен истек'
                
                # Создаем сервисы
                oauth_service = GoogleOAuthService(user)
                calendar_service = GoogleCalendarService(oauth_service)
                
                # Получаем события через API
                service = calendar_service._get_service()
                if not service:
                    return 0, f'Ошибка получения сервиса'
                
                # Формируем запрос
                time_min = start_date.isoformat()
                time_max = sync_end_date.isoformat()
                
                print(f"   📅 Запрашиваем события календаря '{calendar_name}' (ID: {calendar_id})...")
                events_result = service.events().list(
                    calendarId=calendar_id,
                    timeMin=time_min,
                    timeMax=time_max,
                    maxResults=2500,
                    singleEvents=True,
                    orderBy='startTime'
                ).execute()
                
                events = events_result.get('items', [])
                print(f"   📅 Получено {len(events)} событий из календаря '{calendar_name}'")
                
                # Сохраняем события в БД
                synced_count = 0
                for event in events:
                    try:
                        event_id = event.get('id')
                        if not event_id:
                            continue
                        
                        # Парсим время
                        start_time = _parse_event_time_for_sync(event.get('start'))
                        end_time = _parse_event_time_for_sync(event.get('end'))
                        
                        if not start_time or not end_time:
                            continue
                        
                        # Фильтруем по периоду
                        if start_time < start_date or start_time > sync_end_date:
                            continue
                        
                        # Извлекаем участников
                        attendees = []
                        attendee_emails = []
                        for attendee in event.get('attendees', []):
                            attendee_email = attendee.get('email', '').lower()
                            attendees.append({
                                'email': attendee_email,
                                'name': attendee.get('displayName', attendee_email)
                            })
                            attendee_emails.append(attendee_email)
                        
                        # Извлекаем email организатора
                        organizer_email = None
                        if 'organizer' in event:
                            organizer_email = event['organizer'].get('email', '').lower()
                            # Добавляем организатора в список email для проверки
                            if organizer_email and organizer_email not in attendee_emails:
                                attendee_emails.append(organizer_email)
                        
                        # Получаем время обновления из Google
                        google_updated = None
                        if 'updated' in event:
                            try:
                                google_updated = datetime.fromisoformat(event['updated'].replace('Z', '+00:00'))
                            except:
                                pass
                        
                        # Определяем рекрутера для события
                        # Специальная логика: если на встрече нет andrei.golubenko@softnetix.io,
                        # но есть andrei.chernomordin@softnetix.io, то присваиваем событие ему
                        golubenko_email = 'andrei.golubenko@softnetix.io'
                        chernomordin_email = 'andrei.chernomordin@softnetix.io'
                        
                        has_golubenko = golubenko_email in attendee_emails
                        has_chernomordin = chernomordin_email in attendee_emails
                        
                        event_recruiter = user
                        
                        # Приоритетная логика: если нет andrei.golubenko, но есть andrei.chernomordin, присваиваем ему
                        if not has_golubenko and has_chernomordin:
                            chernomordin_user = User.objects.filter(email=chernomordin_email).first()
                            if chernomordin_user:
                                event_recruiter = chernomordin_user
                                print(f"   📌 Событие присвоено {chernomordin_email} (нет {golubenko_email} среди участников)")
                        else:
                            # Стандартная логика определения рекрутера
                            if not user.groups.filter(name='Рекрутер').exists():
                                # Если пользователь не рекрутер, пытаемся найти рекрутера по участникам
                                for attendee_email in attendee_emails:
                                    recruiter_user = User.objects.filter(
                                        email=attendee_email,
                                        groups__name='Рекрутер'
                                    ).first()
                                    if recruiter_user:
                                        event_recruiter = recruiter_user
                                        break
                                
                                # Если не нашли, используем первого рекрутера
                                if not event_recruiter.groups.filter(name='Рекрутер').exists():
                                    first_recruiter = recruiters.first()
                                    if first_recruiter:
                                        event_recruiter = first_recruiter
                        
                        # Сохраняем событие
                        CalendarEvent.objects.update_or_create(
                            event_id=event_id,
                            defaults={
                                'recruiter': event_recruiter,
                                'title': event.get('summary', 'Без названия'),
                                'start_time': start_time,
                                'end_time': end_time,
                                'attendees': attendees,
                                'description': event.get('description', ''),
                                'location': event.get('location', ''),
                                'google_updated_at': google_updated,
                            }
                        )
                        
                        synced_count += 1
                        
                    except Exception as e:
                        print(f"   ❌ Ошибка сохранения события: {e}")
                
                return synced_count, None
                
            except Exception as e:
                error_msg = str(e)
                print(f"   ❌ Ошибка синхронизации календаря '{calendar_name}': {error_msg}")
                import traceback
                print(traceback.format_exc())
                return 0, error_msg
        
        # Синхронизируем календари рекрутеров
        for recruiter in recruiters:
            try:
                print(f"\n👤 Обработка рекрутера: {recruiter.get_full_name() or recruiter.username} ({recruiter.email})")
                
                recruiter_synced, error = sync_calendar_for_user(recruiter, 'primary', f'Рекрутер {recruiter.username}')
                
                if error:
                    skipped_recruiters.append({
                        'name': recruiter.get_full_name() or recruiter.username,
                        'reason': error
                    })
                    print(f"   ⚠️  Пропущен: {error}")
                else:
                    total_synced += recruiter_synced
                    if recruiter_synced > 0:
                        synced_recruiters.append({
                            'name': recruiter.get_full_name() or recruiter.username,
                            'count': recruiter_synced
                        })
                        print(f"   ✅ Синхронизировано {recruiter_synced} событий")
                    else:
                        print(f"   ℹ️  Новых событий не найдено")
                    
            except Exception as e:
                total_errors += 1
                error_msg = str(e)
                print(f"   ❌ Ошибка синхронизации для {recruiter.username}: {error_msg}")
                import traceback
                print(traceback.format_exc())
                skipped_recruiters.append({
                    'name': recruiter.get_full_name() or recruiter.username,
                    'reason': f'Ошибка: {error_msg[:100]}'
                })
        
        # Синхронизируем календари интервьюеров
        # Используем календарь пользователя andrei.golubenko для доступа
        print(f"\n👥 Синхронизация календарей интервьюеров за период {start_date.date()} - {sync_end_date.date()}...")
        interviewer_sync_enabled = request.GET.get('sync_interviewers', 'true').lower() == 'true'
        
        if interviewer_sync_enabled:
            try:
                # Получаем пользователя andrei.golubenko
                golubenko_user = User.objects.filter(email='andrei.golubenko@softnetix.io').first()
                if not golubenko_user:
                    golubenko_user = User.objects.filter(username='andrei.golubenko').first()
                
                if golubenko_user:
                    try:
                        golubenko_oauth = GoogleOAuthAccount.objects.get(user=golubenko_user)
                        if golubenko_oauth.is_token_valid():
                            # Создаем сервисы для доступа к календарям интервьюеров
                            golubenko_oauth_service = GoogleOAuthService(golubenko_user)
                            golubenko_calendar_service = GoogleCalendarService(golubenko_oauth_service)
                            golubenko_service = golubenko_calendar_service._get_service()
                            
                            if golubenko_service:
                                # Получаем всех активных интервьюеров
                                interviewers = Interviewer.objects.filter(is_active=True)
                                print(f"   📋 Найдено {interviewers.count()} активных интервьюеров")
                                
                                # Импортируем функцию извлечения calendar_id
                                from apps.google_oauth.views import _extract_calendar_id_from_link
                                
                                for interviewer in interviewers:
                                    try:
                                        print(f"\n   👤 Обработка интервьюера: {interviewer.get_full_name()} ({interviewer.email})")
                                        
                                        # Определяем calendar_id интервьюера
                                        calendar_id = None
                                        
                                        # Способ 1: Извлекаем из calendar_link
                                        if interviewer.calendar_link:
                                            calendar_id = _extract_calendar_id_from_link(interviewer.calendar_link)
                                            if calendar_id:
                                                print(f"      📅 Извлечен calendar_id из ссылки: {calendar_id}")
                                        
                                        # Способ 2: Ищем календарь по email
                                        if not calendar_id:
                                            try:
                                                calendar = golubenko_calendar_service.get_calendar_by_email(interviewer.email)
                                                if calendar:
                                                    calendar_id = calendar['id']
                                                    print(f"      📅 Найден календарь по email: {calendar_id}")
                                            except Exception as e:
                                                print(f"      ⚠️  Ошибка поиска календаря по email: {e}")
                                        
                                        # Способ 3: Используем email напрямую
                                        if not calendar_id:
                                            calendar_id = interviewer.email
                                            print(f"      📅 Используем email как calendar_id: {calendar_id}")
                                        
                                        if calendar_id:
                                            # Синхронизируем календарь интервьюера
                                            time_min = start_date.isoformat()
                                            time_max = sync_end_date.isoformat()
                                            
                                            print(f"      📅 Запрашиваем события календаря интервьюера (период: {start_date.date()} - {sync_end_date.date()})...")
                                            try:
                                                events_result = golubenko_service.events().list(
                                                    calendarId=calendar_id,
                                                    timeMin=time_min,
                                                    timeMax=time_max,
                                                    maxResults=2500,
                                                    singleEvents=True,
                                                    orderBy='startTime'
                                                ).execute()
                                                
                                                events = events_result.get('items', [])
                                                print(f"      📅 Получено {len(events)} событий из календаря интервьюера")
                                                
                                                # Сохраняем события в БД
                                                interviewer_synced = 0
                                                interviewer_updated = 0
                                                interviewer_created = 0
                                                
                                                for event in events:
                                                    try:
                                                        event_id = event.get('id')
                                                        if not event_id:
                                                            continue
                                                        
                                                        # Парсим время
                                                        start_time = _parse_event_time_for_sync(event.get('start'))
                                                        end_time = _parse_event_time_for_sync(event.get('end'))
                                                        
                                                        if not start_time or not end_time:
                                                            continue
                                                        
                                                        # Фильтруем по периоду
                                                        if start_time < start_date or start_time > sync_end_date:
                                                            continue
                                                        
                                                        # Извлекаем участников
                                                        attendees = []
                                                        for attendee in event.get('attendees', []):
                                                            attendees.append({
                                                                'email': attendee.get('email', ''),
                                                                'name': attendee.get('displayName', attendee.get('email', ''))
                                                            })
                                                        
                                                        # Получаем время обновления из Google
                                                        google_updated = None
                                                        if 'updated' in event:
                                                            try:
                                                                google_updated = datetime.fromisoformat(event['updated'].replace('Z', '+00:00'))
                                                            except:
                                                                pass
                                                        
                                                        # Определяем рекрутера для события интервьюера
                                                        # Пытаемся найти рекрутера по участникам
                                                        event_recruiter = None
                                                        for attendee_email in [a.get('email', '') for a in attendees]:
                                                            recruiter_user = User.objects.filter(
                                                                email=attendee_email,
                                                                groups__name='Рекрутер'
                                                            ).first()
                                                            if recruiter_user:
                                                                event_recruiter = recruiter_user
                                                                break
                                                        
                                                        # Если не нашли, используем первого рекрутера
                                                        if not event_recruiter:
                                                            event_recruiter = recruiters.first()
                                                        
                                                        if not event_recruiter:
                                                            print(f"      ⚠️  Не найден рекрутер для события {event.get('summary', 'Без названия')}. Пропускаем.")
                                                            continue
                                                        
                                                        # Сохраняем событие
                                                        calendar_event, created = CalendarEvent.objects.update_or_create(
                                                            event_id=event_id,
                                                            defaults={
                                                                'recruiter': event_recruiter,
                                                                'title': event.get('summary', 'Без названия'),
                                                                'start_time': start_time,
                                                                'end_time': end_time,
                                                                'attendees': attendees,
                                                                'description': event.get('description', ''),
                                                                'location': event.get('location', ''),
                                                                'google_updated_at': google_updated,
                                                            }
                                                        )
                                                        
                                                        if created:
                                                            interviewer_created += 1
                                                        else:
                                                            interviewer_updated += 1
                                                        
                                                        interviewer_synced += 1
                                                        total_synced += 1
                                                        
                                                    except Exception as e:
                                                        total_errors += 1
                                                        print(f"      ❌ Ошибка сохранения события: {e}")
                                                        import traceback
                                                        print(traceback.format_exc())
                                                
                                                if interviewer_synced > 0:
                                                    synced_recruiters.append({
                                                        'name': f'Интервьюер: {interviewer.get_full_name()} ({interviewer.email})',
                                                        'count': interviewer_synced,
                                                        'created': interviewer_created,
                                                        'updated': interviewer_updated
                                                    })
                                                    print(f"      ✅ Синхронизировано {interviewer_synced} событий (создано: {interviewer_created}, обновлено: {interviewer_updated})")
                                                else:
                                                    print(f"      ℹ️  Новых событий не найдено за указанный период")
                                                    
                                            except Exception as e:
                                                total_errors += 1
                                                error_msg = str(e)
                                                print(f"      ❌ Ошибка запроса событий календаря: {error_msg}")
                                                skipped_recruiters.append({
                                                    'name': f'Интервьюер: {interviewer.get_full_name()}',
                                                    'reason': f'Ошибка запроса: {error_msg[:100]}'
                                                })
                                        else:
                                            skipped_recruiters.append({
                                                'name': f'Интервьюер: {interviewer.get_full_name()}',
                                                'reason': 'Не удалось определить calendar_id'
                                            })
                                            print(f"      ⚠️  Не удалось определить calendar_id")
                                            
                                    except Exception as e:
                                        total_errors += 1
                                        error_msg = str(e)
                                        print(f"      ❌ Ошибка синхронизации для {interviewer.get_full_name()}: {error_msg}")
                                        import traceback
                                        print(traceback.format_exc())
                                        skipped_recruiters.append({
                                            'name': f'Интервьюер: {interviewer.get_full_name()}',
                                            'reason': f'Ошибка: {error_msg[:100]}'
                                        })
                            else:
                                print(f"   ⚠️  Не удалось получить сервис Google Calendar для пользователя andrei.golubenko")
                        else:
                            print(f"   ⚠️  Токен Google OAuth истек для пользователя andrei.golubenko")
                            skipped_recruiters.append({
                                'name': 'Синхронизация интервьюеров',
                                'reason': 'Токен Google OAuth истек для пользователя andrei.golubenko'
                            })
                    except GoogleOAuthAccount.DoesNotExist:
                        print(f"   ⚠️  У пользователя andrei.golubenko нет Google OAuth аккаунта")
                        skipped_recruiters.append({
                            'name': 'Синхронизация интервьюеров',
                            'reason': 'У пользователя andrei.golubenko нет Google OAuth аккаунта'
                        })
                else:
                    print(f"   ⚠️  Пользователь andrei.golubenko не найден")
                    skipped_recruiters.append({
                        'name': 'Синхронизация интервьюеров',
                        'reason': 'Пользователь andrei.golubenko не найден'
                    })
            except Exception as e:
                print(f"   ❌ Ошибка при синхронизации календарей интервьюеров: {e}")
                import traceback
                print(traceback.format_exc())
                skipped_recruiters.append({
                    'name': 'Синхронизация интервьюеров',
                    'reason': f'Ошибка: {str(e)[:100]}'
                })
        else:
            print(f"   ⏭️  Синхронизация интервьюеров пропущена (параметр sync_interviewers=false)")
        
        # Синхронизируем дополнительные календари (например, календарь компании)
        # Для этого используем первого рекрутера с валидным OAuth аккаунтом
        print(f"\n📅 Синхронизация дополнительных календарей...")
        for calendar_info in additional_calendars:
            calendar_id = calendar_info['calendar_id']
            calendar_name = calendar_info['name']
            
            # Находим первого рекрутера с валидным OAuth для доступа к календарю
            user_for_calendar = None
            for recruiter in recruiters:
                try:
                    oauth_account = GoogleOAuthAccount.objects.get(user=recruiter)
                    if oauth_account.is_token_valid():
                        user_for_calendar = recruiter
                        break
                except:
                    continue
            
            if not user_for_calendar:
                print(f"   ⚠️  Не найден рекрутер с валидным OAuth для доступа к календарю '{calendar_name}'. Пропускаем.")
                skipped_recruiters.append({
                    'name': calendar_name,
                    'reason': 'Нет рекрутера с валидным OAuth'
                })
                continue
            
            print(f"\n📅 Обработка календаря: {calendar_name}")
            calendar_synced, error = sync_calendar_for_user(user_for_calendar, calendar_id, calendar_name)
            
            if error:
                skipped_recruiters.append({
                    'name': calendar_name,
                    'reason': error
                })
                print(f"   ⚠️  Пропущен: {error}")
            else:
                total_synced += calendar_synced
                if calendar_synced > 0:
                    synced_recruiters.append({
                        'name': calendar_name,
                        'count': calendar_synced
                    })
                    print(f"   ✅ Синхронизировано {calendar_synced} событий")
                else:
                    print(f"   ℹ️  Новых событий не найдено")
        
        # Подсчитываем статистику по рекрутерам и интервьюерам
        recruiter_count = len([r for r in synced_recruiters if not r['name'].startswith('Интервьюер:')])
        interviewer_count = len([r for r in synced_recruiters if r['name'].startswith('Интервьюер:')])
        
        print(f"\n✅ Синхронизация завершена!")
        print(f"   Период: {start_date.date()} - {sync_end_date.date()}")
        print(f"   Всего синхронизировано: {total_synced} событий")
        print(f"   Ошибок: {total_errors}")
        print(f"   Обработано рекрутеров: {recruiter_count}")
        print(f"   Обработано интервьюеров: {interviewer_count}")
        print(f"   Всего обработано: {len(synced_recruiters)}")
        print(f"   Пропущено: {len(skipped_recruiters)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Синхронизация завершена. Синхронизировано {total_synced} событий (рекрутеры: {recruiter_count}, интервьюеры: {interviewer_count})',
            'synced_count': total_synced,
            'errors': total_errors,
            'start_date': start_date.isoformat(),
            'end_date': sync_end_date.isoformat(),
            'recruiters': synced_recruiters,
            'skipped_recruiters': skipped_recruiters,
            'total_recruiters': recruiters.count(),
            'processed_recruiters': recruiter_count,
            'processed_interviewers': interviewer_count,
            'total_processed': len(synced_recruiters),
            'skipped_count': len(skipped_recruiters),
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)


def _parse_event_time_for_sync(time_data):
    """Парсит время события из формата Google Calendar API"""
    if not time_data:
        return None
    
    if 'dateTime' in time_data:
        try:
            return datetime.fromisoformat(time_data['dateTime'].replace('Z', '+00:00'))
        except:
            pass
    
    if 'date' in time_data:
        try:
            dt = datetime.fromisoformat(time_data['date'])
            if timezone.is_naive(dt):
                return timezone.make_aware(dt)
            return dt
        except:
            pass
    
    return None


@login_required
def export_company_report_excel(request):
    """Экспорт отчета по компании в Excel"""
    period = request.GET.get('period', 'monthly')
    start_date, end_date = parse_date_range(request, period)
    
    generator = ReportGenerator(request.user)
    report_data = generator.generate_company_report(start_date, end_date, period)
    
    # Добавляем даты для экспорта
    report_data['start_date'] = start_date
    report_data['end_date'] = end_date
    report_data['period'] = period
    
    # Получаем детальные данные для таблицы
    from apps.reporting.models import CalendarEvent
    from apps.interviewers.models import Interviewer
    
    events = CalendarEvent.objects.filter(
        start_time__gte=start_date,
        start_time__lte=end_date
    ).select_related('vacancy', 'recruiter').order_by('start_time')
    
    # Группируем события по периоду и рекрутеру
    detailed_data = {}
    for event in events:
        # Определяем период для события
        event_date = event.start_time.date()
        if period == 'daily':
            period_key = event_date.isoformat()
        elif period == 'weekly':
            period_key = f"{event_date.year}-W{event_date.isocalendar()[1]:02d}"
        elif period == 'monthly':
            period_key = f"{event_date.year}-{event_date.month:02d}"
        elif period == 'quarterly':
            quarter = (event_date.month - 1) // 3 + 1
            period_key = f"{event_date.year}-Q{quarter}"
        elif period == 'yearly':
            period_key = str(event_date.year)
        else:
            period_key = event_date.isoformat()
        
        recruiter_id = event.recruiter.id
        recruiter_name = event.recruiter.get_full_name() or event.recruiter.username
        
        # Получаем интервьюеров из участников
        interviewer_emails = set()
        interviewer_names = []
        if event.attendees:
            for attendee in event.attendees:
                if isinstance(attendee, dict):
                    email = attendee.get('email', '').lower()
                elif isinstance(attendee, str):
                    email = attendee.lower()
                else:
                    continue
                
                if email:
                    try:
                        interviewer = Interviewer.objects.filter(
                            email__iexact=email,
                            is_active=True
                        ).first()
                        if interviewer:
                            interviewer_emails.add(email)
                            interviewer_names.append(interviewer.get_full_name())
                    except:
                        pass
        
        # Создаем ключ для группировки
        group_key = (period_key, recruiter_id)
        
        if group_key not in detailed_data:
            detailed_data[group_key] = {
                'period': period_key,
                'recruiter_id': recruiter_id,
                'recruiter_name': recruiter_name,
                'interviewers': set(interviewer_emails),
                'interviewer_names': set(interviewer_names),
                'screenings': 0,
                'interviews': 0,
                'total_time_minutes': 0,
            }
        
        # Обновляем статистику
        if event.event_type == 'screening':
            detailed_data[group_key]['screenings'] += 1
        elif event.event_type == 'interview':
            detailed_data[group_key]['interviews'] += 1
        
        detailed_data[group_key]['total_time_minutes'] += event.duration_minutes or 0
        detailed_data[group_key]['interviewers'].update(interviewer_emails)
        detailed_data[group_key]['interviewer_names'].update(interviewer_names)
    
    # Преобразуем sets в списки для сериализации
    detailed_list = []
    for group_key, data in detailed_data.items():
        detailed_list.append({
            'period': data['period'],
            'recruiter_name': data['recruiter_name'],
            'interviewer_names': ', '.join(sorted(data['interviewer_names'])) if data['interviewer_names'] else '—',
            'screenings': data['screenings'],
            'interviews': data['interviews'],
            'total': data['screenings'] + data['interviews'],
            'total_time_minutes': data['total_time_minutes'],
        })
    
    # Сортируем по периоду и рекрутеру
    detailed_list.sort(key=lambda x: (x['period'], x['recruiter_name']))
    
    report_data['detailed_data'] = detailed_list
    
    exporter = ExcelReportExporter(report_data, 'company', 'Отчет по компании')
    return exporter.export()


@login_required
def export_recruiters_summary_excel(request):
    """Экспорт сводного отчета по рекрутерам в Excel"""
    period = request.GET.get('period', 'monthly')
    start_date, end_date = parse_date_range(request, period)
    
    generator = ReportGenerator(request.user)
    report_data = generator.generate_recruiters_summary_report(start_date, end_date, period)
    
    # Добавляем даты для экспорта
    report_data['start_date'] = start_date
    report_data['end_date'] = end_date
    
    exporter = ExcelReportExporter(report_data, 'recruiters_summary', 'Сводный отчет по рекрутерам')
    return exporter.export()


@login_required
def export_recruiter_report_excel(request, recruiter_id):
    """Экспорт отчета по рекрутеру в Excel"""
    recruiter = get_object_or_404(User, id=recruiter_id)
    period = request.GET.get('period', 'monthly')
    start_date, end_date = parse_date_range(request, period)
    
    interviewer_id = request.GET.get('interviewer_id')
    interviewer_id = int(interviewer_id) if interviewer_id else None
    
    generator = ReportGenerator(request.user)
    report_data = generator.generate_recruiter_report(recruiter, start_date, end_date, period, interviewer_id=interviewer_id)
    
    # Добавляем даты для экспорта
    report_data['start_date'] = start_date
    report_data['end_date'] = end_date
    
    recruiter_name = recruiter.get_full_name() or recruiter.username
    exporter = ExcelReportExporter(report_data, 'recruiter', f'Отчет по рекрутеру: {recruiter_name}')
    return exporter.export()


@login_required
def export_vacancy_report_excel(request, vacancy_id):
    """Экспорт отчета по вакансии в Excel"""
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    period = request.GET.get('period', 'monthly')
    start_date, end_date = parse_date_range(request, period)
    
    recruiter_id = request.GET.get('recruiter_id')
    recruiter_id = int(recruiter_id) if recruiter_id else None
    
    generator = ReportGenerator(request.user)
    report_data = generator.generate_vacancy_report(vacancy, start_date, end_date, period, recruiter_id=recruiter_id)
    
    # Добавляем даты для экспорта
    report_data['start_date'] = start_date
    report_data['end_date'] = end_date
    
    exporter = ExcelReportExporter(report_data, 'vacancy', f'Отчет по вакансии: {vacancy.name}')
    return exporter.export()


@login_required
def export_interviewer_report_excel(request, interviewer_id):
    """Экспорт отчета по интервьюеру в Excel"""
    interviewer = get_object_or_404(Interviewer, id=interviewer_id)
    period = request.GET.get('period', 'monthly')
    start_date, end_date = parse_date_range(request, period)
    
    recruiter_id = request.GET.get('recruiter_id')
    recruiter_id = int(recruiter_id) if recruiter_id else None
    
    generator = ReportGenerator(request.user)
    report_data = generator.generate_interviewer_report(interviewer, start_date, end_date, period, recruiter_id=recruiter_id)
    
    # Добавляем даты для экспорта
    report_data['start_date'] = start_date
    report_data['end_date'] = end_date
    
    interviewer_name = interviewer.get_full_name()
    exporter = ExcelReportExporter(report_data, 'interviewer', f'Отчет по интервьюеру: {interviewer_name}')
    return exporter.export()


@login_required
def export_interviewers_list_excel(request):
    """Экспорт списка интервьюеров со статистикой в Excel"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Получаем параметры фильтров
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    interviewer_ids_str = request.GET.get('interviewer_ids', '')
    chart_data_type = request.GET.get('chart_data_type', 'screenings')
    
    # Парсим диапазон дат
    if start_date_str and end_date_str:
        try:
            start_date_naive = datetime.fromisoformat(start_date_str)
            end_date_naive = datetime.fromisoformat(end_date_str)
            
            if timezone.is_naive(start_date_naive):
                start_date = timezone.make_aware(start_date_naive)
            else:
                start_date = start_date_naive
            
            if timezone.is_naive(end_date_naive):
                end_date = timezone.make_aware(end_date_naive)
            else:
                end_date = end_date_naive
            
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        except:
            end_date = timezone.now()
            start_date = end_date - relativedelta(months=12)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    else:
        end_date = timezone.now()
        start_date = end_date - relativedelta(months=12)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Получаем интервьюеров
    interviewers = Interviewer.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    # Фильтруем по выбранным ID, если указаны
    if interviewer_ids_str:
        try:
            selected_ids = [int(id) for id in interviewer_ids_str.split(',')]
            interviewers = interviewers.filter(id__in=selected_ids)
        except:
            pass
    
    # Получаем все события за период
    all_events = CalendarEvent.objects.filter(
        start_time__gte=start_date,
        start_time__lte=end_date
    ).select_related('vacancy').order_by('start_time')
    
    # Собираем статистику
    interviewer_stats = []
    for interviewer in interviewers:
        interviewer_email_lower = interviewer.email.lower()
        
        interviewer_events = []
        for event in all_events:
            attendees = event.attendees or []
            is_participant = False
            
            for attendee in attendees:
                if isinstance(attendee, dict):
                    attendee_email = attendee.get('email', '').lower()
                    if attendee_email == interviewer_email_lower:
                        is_participant = True
                        break
                elif isinstance(attendee, str):
                    if attendee.lower() == interviewer_email_lower:
                        is_participant = True
                        break
            
            if is_participant:
                interviewer_events.append(event)
        
        screenings = sum(1 for e in interviewer_events if e.event_type == 'screening')
        interviews = sum(1 for e in interviewer_events if e.event_type == 'interview')
        total_time_minutes = sum(e.duration_minutes or 0 for e in interviewer_events)
        
        interviewer_stats.append({
            'interviewer': interviewer,
            'screenings': screenings,
            'interviews': interviews,
            'total': screenings + interviews,
            'total_time_minutes': total_time_minutes,
        })
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по интервьюерам"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    
    # Заголовок
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = "Отчет по интервьюерам"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Информация о фильтрах
    ws.merge_cells(f'A{row}:F{row}')
    filter_info = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    if interviewer_ids_str:
        filter_info += f" | Выбрано интервьюеров: {len(interviewer_stats)}"
    if chart_data_type:
        data_type_labels = {
            'screenings': 'Скрининги',
            'interviews': 'Интервью',
            'total': 'Всего встреч',
            'time': 'Время'
        }
        filter_info += f" | Тип данных графика: {data_type_labels.get(chart_data_type, chart_data_type)}"
    ws[f'A{row}'].value = filter_info
    ws[f'A{row}'].font = Font(italic=True)
    row += 2
    
    # Заголовки таблицы
    headers = ['Интервьюер', 'Email', 'Скрининги', 'Интервью', 'Всего встреч', 'Время']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    row += 1
    
    # Данные таблицы
    table_start_row = row
    for stat in interviewer_stats:
        interviewer = stat['interviewer']
        hours = stat['total_time_minutes'] // 60
        minutes = stat['total_time_minutes'] % 60
        time_str = f"{hours} ч {minutes} мин" if stat['total_time_minutes'] > 0 else "0 ч 0 мин"
        
        data_row = [
            interviewer.get_full_name(),
            interviewer.email,
            stat['screenings'],
            stat['interviews'],
            stat['total'],
            time_str
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col_idx in [3, 4, 5]:  # Числовые столбцы
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
    
    table_end_row = row - 1
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # Добавляем график только если есть данные
    if interviewer_stats:
        row += 2
        chart_title_row = row
        ws.merge_cells(f'A{row}:F{row}')
        chart_title_cell = ws[f'A{row}']
        data_type_labels = {
            'screenings': 'Количество скринингов',
            'interviews': 'Количество интервью',
            'total': 'Общее количество встреч',
            'time': 'Время (часы)'
        }
        chart_title_cell.value = f"График: {data_type_labels.get(chart_data_type, 'Данные')}"
        chart_title_cell.font = Font(bold=True, size=12)
        chart_title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        # Создаем график
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = data_type_labels.get(chart_data_type, 'Данные')
        chart.y_axis.title = data_type_labels.get(chart_data_type, 'Значение')
        chart.x_axis.title = 'Интервьюер'
        
        # Определяем столбец данных в зависимости от типа
        data_col_map = {
            'screenings': 3,
            'interviews': 4,
            'total': 5,
            'time': 6
        }
        data_col = data_col_map.get(chart_data_type, 3)
        
        # Данные для графика
        if chart_data_type == 'time':
            # Для времени создаем отдельный столбец с часами
            chart_data_row = row
            ws.cell(row=row, column=1).value = 'Интервьюер'
            ws.cell(row=row, column=2).value = 'Время (часы)'
            row += 1
            for stat in interviewer_stats:
                ws.cell(row=row, column=1).value = stat['interviewer'].get_full_name()
                ws.cell(row=row, column=2).value = round(stat['total_time_minutes'] / 60, 2)
                row += 1
            chart_data_end_row = row - 1
            
            data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_end_row)
            chart.add_data(data, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_end_row)
            chart.set_categories(cats)
        else:
            data = Reference(ws, min_col=data_col, min_row=table_start_row - 1, max_row=table_end_row)
            chart.add_data(data, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=table_start_row, max_row=table_end_row)
            chart.set_categories(cats)
        
        # Размещаем график
        ws.add_chart(chart, f'A{row}')
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Отчет_по_интервьюерам_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

